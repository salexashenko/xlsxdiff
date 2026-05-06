from __future__ import annotations

import json
import shutil
import sys
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Tuple
from xml.etree import ElementTree as ET

import yaml
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName


REPO_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = REPO_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from workbook_diff import diff_workbooks, write_artifacts
from workbook_diff.reporting import render_llm_summary_markdown
from workbook_diff.utils import write_json


BENCHMARK_DIR = Path(__file__).resolve().parent
CASES_DIR = BENCHMARK_DIR / "cases"
DOCS_REPORT_DIR = REPO_ROOT / "docs" / "examples" / "simple_assumption_report"
FIXED_DATETIME = datetime(2026, 1, 1, 0, 0, 0)
FIXED_ZIP_DATE = (2026, 1, 1, 0, 0, 0)
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_REL)


@dataclass(frozen=True)
class BenchmarkCase:
    slug: str
    title: str
    purpose: str
    expected: str
    builder: Callable[[Path, Path], None]
    config: Dict[str, Any]
    options: Dict[str, Any]


def main() -> None:
    cases = _cases()
    if CASES_DIR.exists():
        shutil.rmtree(CASES_DIR)
    CASES_DIR.mkdir(parents=True, exist_ok=True)

    first_result: Dict[str, Any] | None = None
    first_case_dir: Path | None = None
    for case in cases:
        case_dir = CASES_DIR / case.slug
        case_dir.mkdir(parents=True, exist_ok=True)
        baseline = case_dir / "baseline.xlsx"
        candidate = case_dir / "candidate.xlsx"
        case.builder(baseline, candidate)

        _write_yaml(case_dir / "workbook_diff.yml", {"config": case.config, "options": case.options})
        result = diff_workbooks(baseline, candidate, config=case.config, options=case.options)
        public_result = {key: value for key, value in result.items() if key != "_artifacts"}
        write_json(case_dir / "expected.diff.json", public_result)
        (case_dir / "expected.llm_summary.md").write_text(
            render_llm_summary_markdown(result["llm_summary"]),
            encoding="utf-8",
        )
        (case_dir / "README.md").write_text(_case_readme(case), encoding="utf-8")

        if first_result is None:
            first_result = result
            first_case_dir = case_dir

    if first_result is not None and first_case_dir is not None:
        if DOCS_REPORT_DIR.exists():
            shutil.rmtree(DOCS_REPORT_DIR)
        artifacts = write_artifacts(first_result, DOCS_REPORT_DIR, formats=["html", "json", "md"])
        (DOCS_REPORT_DIR / "README.md").write_text(_sample_report_readme(first_case_dir, artifacts), encoding="utf-8")

    print(f"Wrote {len(cases)} benchmark case(s) to {CASES_DIR}")
    print(f"Wrote checked-in sample report to {DOCS_REPORT_DIR}")


def _cases() -> List[BenchmarkCase]:
    return [
        BenchmarkCase(
            slug="001_simple_assumption",
            title="Simple Assumption Impact",
            purpose="A changed growth-rate assumption flows through an intermediate revenue formula to a final LTV output.",
            expected="`Assumptions!D14` should explain `Summary!G31` with strong confidence and a visible path through `Revenue!G22`.",
            builder=_build_simple_assumption,
            config={
                "outputs": [{"ref": "Summary!G31", "name": "Total LTV"}],
                "assumption_ranges": ["Assumptions!D14:D14"],
            },
            options={},
        ),
        BenchmarkCase(
            slug="002_named_range_change",
            title="Named Range Driver Change",
            purpose="A defined name changes target cells while formulas continue to refer to the same name.",
            expected="`Growth_2027` should act as a non-cell root and explain `Summary!B2` with moderate confidence.",
            builder=_build_named_range_change,
            config={"outputs": [{"ref": "Summary!B2", "name": "2027 Revenue"}]},
            options={},
        ),
        BenchmarkCase(
            slug="003_large_range_membership",
            title="Large Range Membership",
            purpose="A raw-data cell changes inside a large referenced range that is represented as a collapsed range node.",
            expected="`RawData!C582` should reach `Summary!B2` through the virtual `RawData!A1:Z10000` membership edge.",
            builder=_build_large_range_membership,
            config={"outputs": [{"ref": "Summary!B2", "name": "Total raw amount"}], "raw_data_sheets": ["RawData"]},
            options={"max_range_expand_cells": 100},
        ),
        BenchmarkCase(
            slug="004_inserted_modeling_step",
            title="Inserted Modeling Step",
            purpose="A forecast model inserts a GPU expense row, shifting downstream formulas and labels.",
            expected="Semantic alignment should avoid a delete/add storm and group the inserted GPU Expense modeling step.",
            builder=_build_inserted_modeling_step,
            config={"outputs": [{"ref": "Summary!B5", "name": "2026 EBITDA"}, {"ref": "Summary!C5", "name": "2027 EBITDA"}]},
            options={},
        ),
        BenchmarkCase(
            slug="005_manual_calc_mode",
            title="Manual Calculation Caveat",
            purpose="A normal assumption impact happens in a workbook saved with manual calculation mode enabled.",
            expected="The impact path should be detected, but confidence should be downgraded because cached formula values may be stale.",
            builder=_build_manual_calc_mode,
            config={"outputs": [{"ref": "Summary!G31", "name": "Total LTV"}]},
            options={},
        ),
    ]


def _build_simple_assumption(baseline: Path, candidate: Path) -> None:
    _make_assumption_workbook(baseline, growth=0.18, revenue=1180, summary=1180, manual_calc=False)
    _make_assumption_workbook(candidate, growth=0.22, revenue=1220, summary=1220, manual_calc=False)


def _build_named_range_change(baseline: Path, candidate: Path) -> None:
    _make_named_range_driver_workbook(baseline, target_row=13, cached=110)
    _make_named_range_driver_workbook(candidate, target_row=14, cached=120)


def _build_large_range_membership(baseline: Path, candidate: Path) -> None:
    _make_large_range_workbook(baseline, raw_value=10, cached=10)
    _make_large_range_workbook(candidate, raw_value=15, cached=15)


def _build_inserted_modeling_step(baseline: Path, candidate: Path) -> None:
    _make_model_step_baseline(baseline)
    _make_model_step_candidate(candidate)


def _build_manual_calc_mode(baseline: Path, candidate: Path) -> None:
    _make_assumption_workbook(baseline, growth=0.18, revenue=1180, summary=1180, manual_calc=True)
    _make_assumption_workbook(candidate, growth=0.22, revenue=1220, summary=1220, manual_calc=True)


def _make_assumption_workbook(path: Path, growth: float, revenue: int, summary: int, manual_calc: bool) -> None:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = False
    if manual_calc:
        wb.calculation.calcMode = "manual"
    ws = wb.active
    ws.title = "Assumptions"
    ws["A14"] = "Growth Rate"
    ws["D13"] = "2026"
    ws["D14"] = growth
    ws["D14"].number_format = "0.0%"
    rev = wb.create_sheet("Revenue")
    rev["F22"] = 1000
    rev["G21"] = "2027 Revenue"
    rev["G22"] = "=Revenue!F22*(1+Assumptions!D14)"
    customers = wb.create_sheet("Customers")
    customers["G22"] = 1
    summary_ws = wb.create_sheet("Summary")
    summary_ws["A31"] = "Total LTV"
    summary_ws["G31"] = "=Revenue!G22/Customers!G22"
    _save_workbook(wb, path)
    _patch_cached_values(path, {"Revenue": {"G22": revenue}, "Summary": {"G31": summary}})


def _make_named_range_driver_workbook(path: Path, target_row: int, cached: int) -> None:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = False
    ws = wb.active
    ws.title = "Assumptions"
    ws["A13"] = "Old growth"
    ws["D13"] = 0.10
    ws["A14"] = "New growth"
    ws["D14"] = 0.20
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Metric"
    summary["B1"] = "2027E"
    summary["A2"] = "Revenue"
    summary["B2"] = "=100*(1+Growth_2027)"
    wb.defined_names.add(DefinedName("Growth_2027", attr_text=f"'Assumptions'!$D${target_row}"))
    _save_workbook(wb, path)
    _patch_cached_values(path, {"Summary": {"B2": cached}})


def _make_large_range_workbook(path: Path, raw_value: int, cached: int) -> None:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = False
    ws = wb.active
    ws.title = "RawData"
    ws["A1"] = "Header"
    ws["C582"] = raw_value
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Metric"
    summary["B1"] = "Value"
    summary["A2"] = "Total raw amount"
    summary["B2"] = "=SUM(RawData!A1:Z10000)"
    _save_workbook(wb, path)
    _patch_cached_values(path, {"Summary": {"B2": cached}})


def _make_model_step_baseline(path: Path) -> None:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = False
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "2026E"
    ws["C1"] = "2027E"
    ws["A2"] = "Revenue"
    ws["B2"] = 100
    ws["C2"] = 120
    ws["A3"] = "Gross Margin"
    ws["B3"] = "=B2*0.4"
    ws["C3"] = "=C2*0.4"
    ws["A4"] = "EBITDA"
    ws["B4"] = "=B3"
    ws["C4"] = "=C3"
    _save_workbook(wb, path)
    _patch_cached_values(path, {"Summary": {"B3": 40, "C3": 48, "B4": 40, "C4": 48}})


def _make_model_step_candidate(path: Path) -> None:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = False
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "2026E"
    ws["C1"] = "2027E"
    ws["A2"] = "Revenue"
    ws["B2"] = 100
    ws["C2"] = 120
    ws["A3"] = "GPU Expense"
    ws["B3"] = 10
    ws["C3"] = 15
    ws["A4"] = "Gross Margin"
    ws["B4"] = "=B2*0.4"
    ws["C4"] = "=C2*0.4"
    ws["A5"] = "EBITDA"
    ws["B5"] = "=B4-B3"
    ws["C5"] = "=C4-C3"
    _save_workbook(wb, path)
    _patch_cached_values(path, {"Summary": {"B4": 40, "C4": 48, "B5": 30, "C5": 33}})


def _save_workbook(wb: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.properties.creator = "xlsxdiff"
    wb.properties.lastModifiedBy = "xlsxdiff"
    wb.properties.created = FIXED_DATETIME
    wb.properties.modified = FIXED_DATETIME
    wb.save(path)
    _normalize_xlsx_zip(path)


def _normalize_xlsx_zip(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        entries = {info.filename: archive.read(info.filename) for info in archive.infolist()}
    _write_xlsx_entries(path, entries)


def _patch_cached_values(path: Path, values: Dict[str, Dict[str, Any]]) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        entries = {info.filename: archive.read(info.filename) for info in archive.infolist()}
    sheet_paths = _sheet_paths(entries)
    for sheet_name, cell_values in values.items():
        worksheet_path = sheet_paths[sheet_name]
        root = ET.fromstring(entries[worksheet_path])
        for cell_ref, value in cell_values.items():
            cell = root.find(f".//{{{NS_MAIN}}}c[@r='{cell_ref}']")
            if cell is None:
                raise AssertionError(f"Missing cell {sheet_name}!{cell_ref}")
            cell.attrib.pop("t", None)
            value_node = cell.find(f"{{{NS_MAIN}}}v")
            if value_node is None:
                value_node = ET.Element(f"{{{NS_MAIN}}}v")
                formula_node = cell.find(f"{{{NS_MAIN}}}f")
                children = list(cell)
                if formula_node is not None:
                    cell.insert(children.index(formula_node) + 1, value_node)
                else:
                    cell.append(value_node)
            value_node.text = str(value)
        entries[worksheet_path] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    _write_xlsx_entries(path, entries)


def _write_xlsx_entries(path: Path, entries: Dict[str, bytes]) -> None:
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name in sorted(entries):
            info = zipfile.ZipInfo(name, FIXED_ZIP_DATE)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o644 << 16
            archive.writestr(info, entries[name])
    tmp_path.replace(path)


def _sheet_paths(entries: Dict[str, bytes]) -> Dict[str, str]:
    workbook = ET.fromstring(entries["xl/workbook.xml"])
    rels = ET.fromstring(entries["xl/_rels/workbook.xml.rels"])
    rel_targets = {}
    for rel in rels.findall(f"{{{NS_PKG_REL}}}Relationship"):
        target = rel.attrib["Target"]
        if target.startswith("/"):
            target = target.lstrip("/")
        else:
            target = "xl/" + target.lstrip("/")
        rel_targets[rel.attrib["Id"]] = target
    result = {}
    for sheet in workbook.findall(f".//{{{NS_MAIN}}}sheet"):
        name = sheet.attrib["name"]
        rid = sheet.attrib[f"{{{NS_REL}}}id"]
        result[name] = rel_targets[rid]
    return result


def _write_yaml(path: Path, payload: Dict[str, Any]) -> None:
    path.write_text(yaml.safe_dump(payload, sort_keys=False), encoding="utf-8")


def _case_readme(case: BenchmarkCase) -> str:
    return "\n".join(
        [
            f"# {case.title}",
            "",
            f"Purpose: {case.purpose}",
            "",
            f"Expected: {case.expected}",
            "",
            "Files:",
            "",
            "- `baseline.xlsx`",
            "- `candidate.xlsx`",
            "- `workbook_diff.yml`",
            "- `expected.diff.json`",
            "- `expected.llm_summary.md`",
            "",
        ]
    )


def _sample_report_readme(case_dir: Path, artifacts: Dict[str, Path]) -> str:
    rel_case = case_dir.relative_to(REPO_ROOT)
    rel_html = artifacts["html"].relative_to(REPO_ROOT)
    return "\n".join(
        [
            "# Simple Assumption Sample Report",
            "",
            f"This checked-in report is generated from `{rel_case}`.",
            "",
            f"Open `{rel_html}` to view the HTML report.",
            "",
        ]
    )


if __name__ == "__main__":
    main()
