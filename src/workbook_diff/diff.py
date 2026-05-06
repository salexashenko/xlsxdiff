from __future__ import annotations

from collections import defaultdict, deque
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

from .alignment import build_structural_alignment, remap_graph_refs
from .formulas import (
    classify_formula_change,
    expand_reference_cells,
    formula_functions,
    formula_operators,
    numeric_constants,
    normalize_formula,
    parse_formula_references,
)
from .security import check_deadline, enforce_budget, make_deadline, public_resource_budgets, resolve_resource_budgets
from .snapshot import parse_workbook
from .utils import display_scalar, is_number, safe_float, split_cell_id


DIRECT_ROOT_KINDS = {
    "cell_added",
    "cell_deleted",
    "constant_changed",
    "formula_changed",
    "formula_reference_changed",
    "formula_function_changed",
    "formula_operator_changed",
    "formula_constant_changed",
    "formula_named_range_changed",
    "formula_external_reference_changed",
    "formula_anchor_changed_only",
    "formula_text_changed",
    "formula_to_constant",
    "constant_to_formula",
    "defined_name_added",
    "defined_name_deleted",
    "defined_name_changed",
    "table_added",
    "table_deleted",
    "table_range_changed",
    "sheet_added",
    "sheet_deleted",
    "sheet_visibility_changed",
    "macro_presence_changed",
    "external_link_changed",
    "calculation_mode_changed",
}


def diff_workbooks(
    baseline_path: Path,
    candidate_path: Path,
    config: Optional[Dict[str, Any]] = None,
    options: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    config = config or {}
    options = options or {}
    budgets = resolve_resource_budgets(config, options)
    deadline = make_deadline(budgets)
    baseline = parse_workbook(baseline_path, config=config, strict=bool(options.get("strict")), options=options, deadline=deadline)
    check_deadline(deadline, "parsing candidate workbook")
    candidate = parse_workbook(candidate_path, config=config, strict=bool(options.get("strict")), options=options, deadline=deadline)
    if not options.get("include_hidden_sheets", True):
        baseline = _without_hidden_sheets(baseline)
        candidate = _without_hidden_sheets(candidate)
    return diff_snapshots(baseline, candidate, config=config, options=options, deadline=deadline)


def diff_snapshots(
    baseline: Any,
    candidate: Any,
    config: Optional[Dict[str, Any]] = None,
    options: Optional[Dict[str, Any]] = None,
    deadline: Optional[float] = None,
) -> Dict[str, Any]:
    config = config or {}
    options = options or {}
    budgets = resolve_resource_budgets(config, options)
    max_expand = int(_config_value(config, "graph.max_range_expand_cells", options.get("max_range_expand_cells", 1000)))
    baseline_graph_raw = build_dependency_graph(baseline, max_range_expand_cells=max_expand, budgets=budgets, deadline=deadline)
    candidate_graph = build_dependency_graph(candidate, max_range_expand_cells=max_expand, budgets=budgets, deadline=deadline)
    check_deadline(deadline, "building structural alignment")
    alignment = build_structural_alignment(baseline, candidate)
    baseline_graph = remap_graph_refs(baseline_graph_raw, alignment["old_to_new"], edge_id_prefix="baseline_")

    changes: List[Dict[str, Any]] = []
    diagnostics: List[Dict[str, Any]] = []
    diagnostics.extend(_with_workbook_ref("baseline", baseline.diagnostics))
    diagnostics.extend(_with_workbook_ref("candidate", candidate.diagnostics))
    diagnostics.extend(_alignment_diagnostics(alignment))

    change_counter = 1
    change_counter = _diff_sheets(baseline, candidate, changes, change_counter)
    change_counter = _diff_defined_names(baseline, candidate, changes, change_counter)
    change_counter = _diff_tables(baseline, candidate, changes, change_counter)
    change_counter = _diff_workbook_file_info(baseline, candidate, changes, change_counter)
    change_counter = _diff_cells(baseline, candidate, alignment, changes, change_counter, options)

    preliminary_groups: List[Dict[str, Any]] = []
    _add_named_range_rebound_groups(changes, preliminary_groups, 1, baseline, candidate)
    suppressed_root_change_ids = _suppressed_change_ids_from_groups(preliminary_groups)
    root_refs: Set[str] = set()
    root_change_ids_by_ref = defaultdict(list)
    for change in changes:
        if change["directness"] != "direct":
            continue
        if change.get("id") in suppressed_root_change_ids:
            continue
        for root_ref in _change_graph_root_refs(change):
            root_refs.add(root_ref)
            root_change_ids_by_ref[root_ref].append(change["id"])
    changed_cell_refs = {ref for ref in root_refs if _is_cell_ref(ref)}
    add_virtual_range_membership_edges(baseline_graph, changed_cell_refs, budgets=budgets)
    add_virtual_range_membership_edges(candidate_graph, changed_cell_refs, budgets=budgets)

    check_deadline(deadline, "tracing reachable nodes")
    old_reachable = downstream_nodes(root_refs, baseline_graph)
    new_reachable = downstream_nodes(root_refs, candidate_graph)
    all_reachable = old_reachable | new_reachable

    for change in changes:
        if change["kind"] != "cached_value_changed":
            continue
        if change["object_ref"] in all_reachable:
            change["directness"] = "propagated"
            change["confidence"] = min(change["confidence"], 0.85)
        else:
            change["directness"] = "unexplained"
            change["confidence"] = min(change["confidence"], 0.45)
            change["warnings"].append(
                {
                    "severity": "warning",
                    "code": "UNEXPLAINED_VALUE_CHANGE",
                    "message": "Cached value changed without a detected upstream direct change.",
                    "object_ref": change["object_ref"],
                }
            )

    for change in changes:
        change["materiality_score"] = materiality_score(change, candidate_graph, config)

    changes.sort(key=lambda item: (-item["materiality_score"], item["id"]))
    grouped_changes = group_changes(changes, baseline, candidate)
    impacted_outputs = build_impacted_outputs(changes, baseline, candidate, baseline_graph, candidate_graph, root_change_ids_by_ref, config)
    final_outputs, impacted_intermediates = split_impacted_outputs(impacted_outputs, config)
    impact_dag = build_change_impact_dag(changes, impacted_outputs, candidate_graph, baseline_graph)

    diagnostics.append(
        {
            "severity": "info",
            "code": "CACHED_VALUE_MODE",
            "message": "This report uses cached formula results. Numeric output deltas assume both workbooks were saved after recalculation.",
        }
    )
    diagnostics.extend(_result_diagnostics(changes))

    summary = build_summary(changes, impacted_outputs, baseline, candidate, diagnostics, alignment)
    llm_summary = build_llm_summary(changes, impacted_outputs, grouped_changes, summary, diagnostics, config)
    summary["one_sentence_summary"] = llm_summary["one_sentence_summary"]
    direct_changes = _top_direct_changes(changes, grouped_changes)
    unexplained_changes = [change for change in changes if change["directness"] == "unexplained"]

    return {
        "schema_version": "0.1",
        "resource_budgets": public_resource_budgets(budgets),
        "baseline": baseline.file,
        "candidate": candidate.file,
        "summary": summary,
        "llm_summary": llm_summary,
        "changes": changes,
        "grouped_changes": grouped_changes,
        "dependency_graph_stats": {
            "baseline_node_count": len(baseline_graph_raw["nodes"]),
            "baseline_edge_count": len(baseline_graph_raw["edges"]),
            "aligned_baseline_node_count": len(baseline_graph["nodes"]),
            "aligned_baseline_edge_count": len(baseline_graph["edges"]),
            "candidate_node_count": len(candidate_graph["nodes"]),
            "candidate_edge_count": len(candidate_graph["edges"]),
        },
        "structural_alignment": _public_alignment(alignment),
        "change_impact_dag": impact_dag,
        "top_direct_changes": direct_changes[:25],
        "top_impacted_outputs": impacted_outputs[:25],
        "final_outputs": final_outputs[:25],
        "impacted_intermediates": impacted_intermediates[:50],
        "unexplained_changes": unexplained_changes[:50],
        "diagnostics": diagnostics,
        "_artifacts": {
            "baseline_snapshot": baseline.to_dict(),
            "candidate_snapshot": candidate.to_dict(),
            "structural_alignment": alignment,
            "candidate_graph": candidate_graph,
        },
    }


def build_dependency_graph(
    snapshot: Any,
    max_range_expand_cells: int = 1000,
    budgets: Optional[Dict[str, Any]] = None,
    deadline: Optional[float] = None,
) -> Dict[str, Any]:
    budgets = budgets or {}
    nodes: Dict[str, Dict[str, Any]] = {}
    edges: List[Dict[str, Any]] = []
    edge_seen: Set[Tuple[str, str, str, str]] = set()
    edge_counter = 1

    for ref, cell in snapshot.cells.items():
        nodes[ref] = _dependency_node(ref, cell)
    _enforce_graph_budgets(nodes, edges, budgets, "initial graph nodes")

    edge_counter = _add_defined_name_graph_nodes(snapshot, nodes, edges, edge_seen, edge_counter, max_range_expand_cells)
    edge_counter = _add_table_graph_nodes(snapshot, nodes, edges, edge_seen, edge_counter, max_range_expand_cells)
    _enforce_graph_budgets(nodes, edges, budgets, "defined names and tables")

    for formula_ref, cell in snapshot.cells.items():
        check_deadline(deadline, f"building graph dependencies for {formula_ref}")
        if cell.kind != "formula" or not cell.formula:
            continue
        for precedent in cell.formula.precedents:
            target = precedent["target"]
            edge_type = _edge_type(precedent["kind"])
            expanded_cells, expanded = expand_reference_cells(target, max_range_expand_cells)
            if expanded and expanded_cells:
                for precedent_ref in expanded_cells:
                    if precedent_ref not in nodes:
                        nodes[precedent_ref] = _placeholder_node(precedent_ref)
                    edge_counter = _add_dependency_edge(
                        edges,
                        edge_seen,
                        edge_counter,
                        precedent_ref,
                        formula_ref,
                        edge_type,
                        precedent.get("raw"),
                        precedent.get("confidence", 1.0),
                        formula_ref=formula_ref,
                    )
            else:
                if target not in nodes:
                    nodes[target] = _range_or_external_node(target, precedent)
                edge_counter = _add_dependency_edge(
                    edges,
                    edge_seen,
                    edge_counter,
                    target,
                    formula_ref,
                    edge_type,
                    precedent.get("raw"),
                    precedent.get("confidence", 1.0),
                    formula_ref=formula_ref,
                )
            _enforce_graph_budgets(nodes, edges, budgets, formula_ref)

    return {
        "nodes": nodes,
        "edges": edges,
        "adjacency": _adjacency(edges),
        "reverse_adjacency": _reverse_adjacency(edges),
    }


def _enforce_graph_budgets(nodes: Dict[str, Dict[str, Any]], edges: Sequence[Dict[str, Any]], budgets: Dict[str, Any], subject: str) -> None:
    enforce_budget("max_graph_nodes", len(nodes), budgets, f"Dependency graph nodes after {subject}")
    enforce_budget("max_graph_edges", len(edges), budgets, f"Dependency graph edges after {subject}")


def _add_defined_name_graph_nodes(
    snapshot: Any,
    nodes: Dict[str, Dict[str, Any]],
    edges: List[Dict[str, Any]],
    edge_seen: Set[Tuple[str, str, str, str]],
    edge_counter: int,
    max_range_expand_cells: int,
) -> int:
    for defined_name in snapshot.defined_names:
        name = defined_name.get("name")
        target = defined_name.get("ref")
        if not name or not target:
            continue
        node_id = _name_node_id(name)
        nodes[node_id] = {
            "id": node_id,
            "ref": name,
            "node_type": "defined_name",
            "label": name,
            "semantic_role": "unknown",
            "changes": [],
            "materiality_score": 0,
            "confidence": 0.85,
        }
        edge_counter = _add_target_edges(
            node_id,
            target,
            "defined_name_target",
            defined_name.get("raw") or target,
            0.85,
            nodes,
            edges,
            edge_seen,
            edge_counter,
            max_range_expand_cells,
        )
    return edge_counter


def _add_table_graph_nodes(
    snapshot: Any,
    nodes: Dict[str, Dict[str, Any]],
    edges: List[Dict[str, Any]],
    edge_seen: Set[Tuple[str, str, str, str]],
    edge_counter: int,
    max_range_expand_cells: int,
) -> int:
    for table in snapshot.tables:
        name = table.get("name")
        target = table.get("ref")
        if not name or not target:
            continue
        node_id = _table_node_id(name)
        nodes[node_id] = {
            "id": node_id,
            "ref": name,
            "node_type": "table",
            "label": name,
            "semantic_role": "raw_data",
            "sheet_name": table.get("sheet_name"),
            "changes": [],
            "materiality_score": 0,
            "confidence": 0.85,
        }
        edge_counter = _add_target_edges(
            node_id,
            target,
            "table_range_target",
            target,
            0.8,
            nodes,
            edges,
            edge_seen,
            edge_counter,
            max_range_expand_cells,
        )
    return edge_counter


def _add_target_edges(
    source_ref: str,
    target: str,
    edge_type: str,
    evidence: Any,
    confidence: float,
    nodes: Dict[str, Dict[str, Any]],
    edges: List[Dict[str, Any]],
    edge_seen: Set[Tuple[str, str, str, str]],
    edge_counter: int,
    max_range_expand_cells: int,
) -> int:
    expanded_cells, expanded = expand_reference_cells(target, max_range_expand_cells)
    if expanded and expanded_cells:
        for target_ref in expanded_cells:
            if target_ref not in nodes:
                nodes[target_ref] = _placeholder_node(target_ref)
            edge_counter = _add_dependency_edge(edges, edge_seen, edge_counter, source_ref, target_ref, edge_type, evidence, confidence)
        return edge_counter
    if target not in nodes:
        nodes[target] = _range_or_external_node(target, {"kind": "range", "confidence": confidence})
    return _add_dependency_edge(edges, edge_seen, edge_counter, source_ref, target, edge_type, evidence, confidence)


def _add_dependency_edge(
    edges: List[Dict[str, Any]],
    edge_seen: Set[Tuple[str, str, str, str]],
    edge_counter: int,
    from_ref: str,
    to_ref: str,
    edge_type: str,
    evidence: Any,
    confidence: float,
    formula_ref: Optional[str] = None,
) -> int:
    evidence_text = "" if evidence is None else str(evidence)
    key = (from_ref, to_ref, edge_type, evidence_text)
    if key in edge_seen:
        return edge_counter
    edge_seen.add(key)
    edges.append(
        {
            "id": f"e{edge_counter:06d}",
            "from": from_ref,
            "to": to_ref,
            "edge_type": edge_type,
            "formula_ref": formula_ref,
            "evidence": evidence,
            "confidence": confidence,
        }
    )
    return edge_counter + 1


def add_virtual_range_membership_edges(graph: Dict[str, Any], changed_cell_refs: Iterable[str], budgets: Optional[Dict[str, Any]] = None) -> None:
    budgets = budgets or {}
    range_index = _range_node_index(graph)
    if not range_index:
        return
    edges = graph["edges"]
    edge_seen = {(edge["from"], edge["to"], edge.get("edge_type", ""), edge.get("evidence", "")) for edge in edges}
    edge_counter = len(edges) + 1
    virtual_edge_count = 0
    membership_counts: Dict[str, int] = defaultdict(int)
    for cell_ref in sorted(changed_cell_refs, key=_ref_sort_key):
        if cell_ref not in graph["nodes"]:
            graph["nodes"][cell_ref] = _placeholder_node(cell_ref)
        parsed_cell = _parse_cell_ref(cell_ref)
        if parsed_cell is None:
            continue
        cell_sheet, row, col = parsed_cell
        for min_row, max_row, min_col, max_col, range_ref in range_index.get(cell_sheet, []):
            if not (min_row <= row <= max_row and min_col <= col <= max_col):
                continue
            membership_counts[range_ref] += 1
            key = (cell_ref, range_ref, "range_membership", "changed cell inside referenced range")
            if key in edge_seen:
                continue
            edge_seen.add(key)
            virtual_edge_count += 1
            enforce_budget("max_virtual_membership_edges", virtual_edge_count, budgets, "Virtual range membership edges")
            edges.append(
                {
                    "id": f"vrange_{edge_counter:06d}",
                    "from": cell_ref,
                    "to": range_ref,
                    "edge_type": "range_membership",
                    "formula_ref": None,
                    "evidence": "changed cell inside referenced range",
                    "confidence": 0.9,
                }
            )
            edge_counter += 1
            _enforce_graph_budgets(graph["nodes"], edges, budgets, "virtual range membership")
    graph["range_membership_summary"] = [
        {"range_ref": range_ref, "changed_cell_count": count}
        for range_ref, count in sorted(membership_counts.items(), key=lambda item: _ref_sort_key(item[0]))
    ]
    graph["adjacency"] = _adjacency(edges)
    graph["reverse_adjacency"] = _reverse_adjacency(edges)


def _range_node_index(graph: Dict[str, Any]) -> Dict[str, List[Tuple[int, int, int, int, str]]]:
    index: Dict[str, List[Tuple[int, int, int, int, str]]] = defaultdict(list)
    for node_id, node in graph.get("nodes", {}).items():
        if node.get("node_type") != "range" or not _is_range_ref(node_id):
            continue
        bounds = _parse_range_ref(node_id)
        if bounds is None:
            continue
        sheet, min_row, max_row, min_col, max_col = bounds
        index[sheet].append((min_row, max_row, min_col, max_col, node_id))
    for ranges in index.values():
        ranges.sort(key=lambda item: (item[0], item[1], item[2], item[3], item[4]))
    return index


def downstream_nodes(root_refs: Iterable[str], graph: Dict[str, Any]) -> Set[str]:
    seen: Set[str] = set()
    queue = deque(root_refs)
    while queue:
        node = queue.popleft()
        for neighbor in graph["adjacency"].get(node, []):
            if neighbor in seen:
                continue
            seen.add(neighbor)
            queue.append(neighbor)
    return seen


def representative_path(root: str, target: str, graph: Dict[str, Any], max_depth: int = 20) -> Optional[Dict[str, Any]]:
    if root == target:
        return {"nodes": [root], "edge_ids": [], "confidence": 1.0}
    queue = deque([(root, [root], [], 1.0)])
    visited = {root}
    edge_lookup = defaultdict(list)
    for edge in graph["edges"]:
        edge_lookup[(edge["from"], edge["to"])].append(edge)
    while queue:
        node, path, edge_ids, confidence = queue.popleft()
        if len(path) > max_depth:
            continue
        for neighbor in graph["adjacency"].get(node, []):
            if neighbor in path:
                continue
            best_edge = edge_lookup[(node, neighbor)][0]
            next_edge_ids = edge_ids + [best_edge["id"]]
            next_confidence = min(confidence, best_edge.get("confidence", 1.0))
            if neighbor == target:
                return {"nodes": path + [neighbor], "edge_ids": next_edge_ids, "confidence": next_confidence}
            if neighbor not in visited:
                visited.add(neighbor)
                queue.append((neighbor, path + [neighbor], next_edge_ids, next_confidence))
    return None


def _path_display_nodes(
    node_ids: Sequence[str],
    candidate_graph: Dict[str, Any],
    baseline_graph: Dict[str, Any],
    label_overrides: Optional[Dict[str, str]] = None,
) -> List[Dict[str, Any]]:
    display_nodes: List[Dict[str, Any]] = []
    label_overrides = label_overrides or {}
    for node_id in node_ids:
        node = candidate_graph["nodes"].get(node_id) or baseline_graph["nodes"].get(node_id) or _placeholder_node_dict(node_id)
        ref = node.get("ref") or node_id
        display_nodes.append(
            {
                "id": node_id,
                "ref": ref,
                "label": label_overrides.get(ref) or label_overrides.get(node_id) or node.get("label") or ref,
                "node_type": node.get("node_type", "unknown"),
                "semantic_role": node.get("semantic_role", "unknown"),
            }
        )
    return display_nodes


def materiality_score(change: Dict[str, Any], graph: Dict[str, Any], config: Dict[str, Any]) -> float:
    score = 0.0
    role = change.get("semantic_role")
    if role == "output":
        score += 50
    elif role == "assumption":
        score += 40
    elif role == "raw_data":
        score += 12
    if change["kind"].startswith("formula_") or change["kind"] in {"formula_changed", "constant_to_formula", "formula_to_constant"}:
        score += 35
    if change.get("directness") == "propagated":
        score += 20
    if change.get("directness") == "unexplained":
        score += 18
    reachable: Set[str] = set()
    for root_ref in _change_graph_root_refs(change):
        reachable.update(downstream_nodes([root_ref], graph))
    score += min(len(reachable) * 0.5, 30)
    if change.get("delta"):
        delta = change["delta"]
        if delta.get("relative_delta") is not None:
            score += min(abs(delta["relative_delta"]) * 100, 25)
        score += min(abs(delta.get("absolute_delta", 0)) / 1000, 20)
    if change.get("visibility_context", {}).get("sheet_visible", True):
        score += 10
    else:
        score += 3
    if change.get("confidence", 1.0) < 0.6:
        score -= 20
    if change["kind"] == "style_changed":
        score -= 50
    return round(score, 3)


def group_changes(changes: Sequence[Dict[str, Any]], baseline: Any = None, candidate: Any = None) -> List[Dict[str, Any]]:
    groups: List[Dict[str, Any]] = []
    counter = _add_named_range_rebound_groups(changes, groups, 1, baseline, candidate)
    counter = _add_modeling_step_groups(changes, groups, counter)
    suppressed_ids = _suppressed_change_ids_from_groups(groups)
    buckets: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for change in changes:
        if change.get("id") in suppressed_ids:
            continue
        if change["scope"] != "cell":
            continue
        sheet = change.get("sheet_name") or split_cell_id(change["object_ref"])[0]
        family = _change_family(change)
        buckets[(sheet, family, change.get("semantic_role", "unknown"))].append(change)

    for (sheet, family, role), bucket in buckets.items():
        if len(bucket) < 3:
            continue
        refs = [change["object_ref"] for change in bucket]
        range_ref = _bounding_range(refs)
        if role == "raw_data" and len(bucket) >= 20:
            group_type = "raw_data_update"
        elif family == "formula":
            group_type = "formula_block"
        elif role == "assumption":
            group_type = "assumption_block"
        else:
            group_type = "contiguous_range"
        groups.append(
            {
                "id": f"grp_{counter:03d}",
                "group_type": group_type,
                "range_ref": range_ref,
                "label": f"{len(bucket)} {family} changes on {sheet}",
                "change_count": len(bucket),
                "representative_changes": [change["id"] for change in bucket[:5]],
                "summary": f"{len(bucket)} {family} changes detected in {range_ref}.",
                "materiality_score": round(sum(change.get("materiality_score", 0) for change in bucket) / len(bucket), 3),
            }
        )
        counter += 1
    groups.sort(key=lambda item: (-item["materiality_score"], item["id"]))
    return groups


def _top_direct_changes(changes: Sequence[Dict[str, Any]], grouped_changes: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    suppressed_ids = _suppressed_change_ids_from_groups(grouped_changes)
    return [change for change in changes if change.get("directness") == "direct" and change.get("id") not in suppressed_ids][:25]


def _suppressed_change_ids_from_groups(groups: Sequence[Dict[str, Any]]) -> Set[str]:
    suppressed: Set[str] = set()
    for group in groups:
        suppressed.update(str(change_id) for change_id in group.get("suppressed_change_ids", []) if change_id)
    return suppressed


def _add_named_range_rebound_groups(
    changes: Sequence[Dict[str, Any]],
    groups: List[Dict[str, Any]],
    counter: int,
    baseline: Any,
    candidate: Any,
) -> int:
    for name_change in changes:
        if name_change.get("kind") != "defined_name_changed":
            continue
        old_target = name_change.get("old")
        new_target = name_change.get("new")
        if not isinstance(old_target, str) or not isinstance(new_target, str):
            continue
        related = [name_change]
        suppressed: List[str] = []
        targets = {old_target, new_target}
        for change in changes:
            if change.get("id") == name_change.get("id") or change.get("scope") != "cell":
                continue
            refs = {ref for ref in [change.get("object_ref"), change.get("old_ref"), change.get("new_ref")] if ref}
            if refs & targets:
                related.append(change)
                suppressed.append(change["id"])

        old_value = _cell_value_dict(baseline, old_target)
        new_value = _cell_value_dict(candidate, new_target)
        delta = numeric_delta(old_value, new_value)
        old_display = (old_value or {}).get("display_value", "")
        new_display = (new_value or {}).get("display_value", "")
        groups.append(
            {
                "id": f"grp_{counter:03d}",
                "group_type": "named_range_rebound",
                "range_ref": f"{old_target},{new_target}",
                "label": f"Named range rebound: {name_change['object_ref']}",
                "change_count": len(related),
                "representative_changes": [change["id"] for change in related[:8]],
                "primary_change_id": name_change["id"],
                "suppressed_change_ids": sorted(set(suppressed)),
                "name": name_change["object_ref"],
                "old_target": old_target,
                "new_target": new_target,
                "effective_old": old_display,
                "effective_new": new_display,
                "delta": delta,
                "summary": _named_range_rebound_summary(name_change["object_ref"], old_target, new_target, old_display, new_display, delta),
                "materiality_score": max(90.0, max((change.get("materiality_score", 0) for change in related), default=0)),
            }
        )
        counter += 1
    return counter


def _cell_value_dict(snapshot: Any, ref: str) -> Optional[Dict[str, Any]]:
    if snapshot is None or not ref:
        return None
    cell = getattr(snapshot, "cells", {}).get(ref)
    if not cell:
        return None
    return cell.value.to_dict()


def _named_range_rebound_summary(name: str, old_target: str, new_target: str, old_display: str, new_display: str, delta: Optional[Dict[str, Any]]) -> str:
    value_phrase = ""
    if old_display or new_display:
        value_phrase = f" Effective value changed from {old_display or 'blank'} to {new_display or 'blank'}"
        delta_text = _delta_text({"delta": delta}) if delta else ""
        if delta_text:
            value_phrase += f" ({delta_text})"
        value_phrase += "."
    return f"Named range '{name}' was rebound from {old_target} to {new_target}.{value_phrase}"


def _add_modeling_step_groups(changes: Sequence[Dict[str, Any]], groups: List[Dict[str, Any]], counter: int) -> int:
    buckets: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for change in changes:
        if change.get("scope") != "cell" or change.get("kind") not in {"cell_added", "cell_deleted"}:
            continue
        identity = change.get("semantic_identity") or {}
        row_headers = identity.get("row_headers") or []
        label = row_headers[0] if row_headers else _best_label_from_change(change)
        if not label:
            continue
        sheet = change.get("sheet_name") or split_cell_id(change["object_ref"])[0]
        buckets[(sheet, change["kind"], label)].append(change)

    for (sheet, kind, label), bucket in sorted(buckets.items(), key=lambda item: (item[0], _ref_sort_key(item[1][0]["object_ref"]))):
        if len(bucket) < 2:
            continue
        inserted = kind == "cell_added"
        group_type = "modeling_step_inserted" if inserted else "modeling_step_deleted"
        verb = "Inserted" if inserted else "Deleted"
        groups.append(
            {
                "id": f"grp_{counter:03d}",
                "group_type": group_type,
                "range_ref": _bounding_range([change["object_ref"] for change in bucket]),
                "label": f"{verb} modeling step: {label}",
                "change_count": len(bucket),
                "representative_changes": [change["id"] for change in bucket[:5]],
                "summary": f"{verb} modeling step '{label}' on {sheet}.",
                "materiality_score": round(sum(change.get("materiality_score", 0) for change in bucket) / len(bucket), 3),
            }
        )
        counter += 1
    return counter


def build_impacted_outputs(
    changes: Sequence[Dict[str, Any]],
    baseline: Any,
    candidate: Any,
    baseline_graph: Dict[str, Any],
    candidate_graph: Dict[str, Any],
    root_change_ids_by_ref: Dict[str, List[str]],
    config: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    changes_by_id = {change["id"]: change for change in changes}
    root_refs = sorted(root_change_ids_by_ref)
    configured_output_names = _configured_output_names(config or {})
    outputs: List[Dict[str, Any]] = []
    for change in changes:
        if change["scope"] != "cell":
            continue
        if change.get("semantic_role") != "output" and change.get("directness") not in {"propagated", "unexplained"}:
            continue
        if change["kind"] not in {"cached_value_changed", "formula_changed", "formula_reference_changed", "formula_function_changed", "formula_operator_changed", "formula_constant_changed", "formula_named_range_changed", "formula_external_reference_changed", "formula_text_changed", "constant_to_formula", "formula_to_constant"}:
            continue
        ref = change["object_ref"]
        upstream_ids: List[str] = []
        paths: List[Dict[str, Any]] = []
        path_confidences: List[float] = []
        for root in root_refs:
            path = representative_path(root, ref, candidate_graph) or representative_path(root, ref, baseline_graph)
            if not path:
                continue
            upstream_ids.extend(root_change_ids_by_ref.get(root, []))
            path_id = f"path_{len(paths) + 1:03d}_{len(outputs) + 1:03d}"
            paths.append(
                {
                    "id": path_id,
                    "from": root,
                    "to": ref,
                    "nodes": path["nodes"],
                    "display_nodes": _path_display_nodes(path["nodes"], candidate_graph, baseline_graph, configured_output_names),
                    "edges": path["edge_ids"],
                    "collapsed": False,
                    "confidence": path["confidence"],
                }
            )
            path_confidences.append(path["confidence"])
        upstream_ids = sorted(set(upstream_ids))
        if change["directness"] == "unexplained" or not upstream_ids:
            strength = "unexplained"
        elif len(upstream_ids) == 1 and min(path_confidences or [1.0]) >= 0.85 and change["kind"] == "cached_value_changed":
            strength = "strong"
        elif min(path_confidences or [1.0]) >= 0.65:
            strength = "moderate"
        else:
            strength = "weak"
        strength, confidence_factors, dependency_confidence, value_delta_confidence = degrade_explanation_strength(
            strength,
            paths,
            upstream_ids,
            change,
            changes_by_id,
            baseline,
            candidate,
            baseline_graph,
            candidate_graph,
        )

        old_ref = change.get("old_ref") or ref
        new_ref = change.get("new_ref") or ref
        old_cell = baseline.cells.get(old_ref)
        new_cell = candidate.cells.get(new_ref)
        outputs.append(
            {
                "ref": ref,
                "old_ref": old_ref,
                "new_ref": new_ref,
                "address_changed": bool(change.get("address_changed")),
                "change_id": change.get("id"),
                "change_kind": change.get("kind"),
                "semantic_id": change.get("semantic_id"),
                "label": configured_output_names.get(ref) or _best_label_from_change(change),
                "old_value": old_cell.value.to_dict() if old_cell else None,
                "new_value": new_cell.value.to_dict() if new_cell else None,
                "delta": change.get("delta"),
                "upstream_change_ids": upstream_ids,
                "representative_paths": paths[:5],
                "explanation_strength": strength,
                "dependency_confidence": dependency_confidence,
                "value_delta_confidence": value_delta_confidence,
                "confidence_factors": confidence_factors,
                "explanation": _output_explanation(ref, change, upstream_ids, strength),
                "materiality_score": change.get("materiality_score", 0),
                "semantic_role": change.get("semantic_role", "unknown"),
            }
        )
    outputs.sort(key=lambda item: (-item["materiality_score"], item["ref"]))
    return outputs


def split_impacted_outputs(
    impacted_outputs: Sequence[Dict[str, Any]],
    config: Optional[Dict[str, Any]] = None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    ranked_outputs = _rank_outputs_for_summary(impacted_outputs, config or {})
    configured_refs = _configured_output_refs(config or {})
    intermediate_refs = _path_intermediate_output_refs(ranked_outputs)
    final_outputs: List[Dict[str, Any]] = []
    intermediates: List[Dict[str, Any]] = []
    for output in ranked_outputs:
        if _is_final_output(output, configured_refs, intermediate_refs):
            final_outputs.append(output)
        else:
            intermediates.append(output)
    if not final_outputs and ranked_outputs:
        final_outputs = [ranked_outputs[0]]
        intermediates = list(ranked_outputs[1:])
    return final_outputs, intermediates


def degrade_explanation_strength(
    strength: str,
    paths: Sequence[Dict[str, Any]],
    upstream_ids: Sequence[str],
    output_change: Dict[str, Any],
    changes_by_id: Dict[str, Dict[str, Any]],
    baseline: Any,
    candidate: Any,
    baseline_graph: Dict[str, Any],
    candidate_graph: Dict[str, Any],
) -> Tuple[str, List[Dict[str, Any]], str, str]:
    factors: List[Dict[str, Any]] = []

    def add_factor(code: str, message: str, cap: str, area: str = "explanation") -> None:
        factors.append({"code": code, "message": message, "strength_cap": cap, "confidence_area": area})

    for workbook_name, snapshot in [("baseline", baseline), ("candidate", candidate)]:
        if snapshot.file.get("calc_mode") == "manual":
            add_factor("manual_calculation_mode", f"{workbook_name} workbook uses manual calculation mode.", "moderate", "value_delta")
        if snapshot.file.get("full_calc_on_load"):
            add_factor("full_calc_on_load", f"{workbook_name} workbook requests full recalculation on load.", "moderate", "value_delta")

    upstream_changes = [changes_by_id[item] for item in upstream_ids if item in changes_by_id]
    if len(upstream_changes) > 1:
        add_factor("multiple_upstream_roots", "Multiple upstream direct changes converge on this output.", "moderate")
    if any(change.get("scope") != "cell" for change in upstream_changes):
        add_factor("non_cell_root", "At least one upstream root is a named range, table, or workbook object.", "moderate")
    if output_change.get("kind", "").startswith("formula_") and _delta_is_nonzero(output_change.get("delta")):
        add_factor("formula_and_value_changed", "The output formula and cached value both changed.", "moderate", "value_delta")
    if output_change.get("kind", "").startswith("formula_") and not _delta_is_nonzero(output_change.get("delta")):
        add_factor("formula_changed_cache_unchanged", "The output formula changed but its cached value did not change.", "moderate", "value_delta")

    output_ref = output_change.get("object_ref")
    old_ref = output_change.get("old_ref") or output_ref
    new_ref = output_change.get("new_ref") or output_ref
    for workbook_name, graph, node_id in [("baseline", baseline_graph, old_ref), ("candidate", candidate_graph, new_ref)]:
        node = graph.get("nodes", {}).get(node_id)
        if _node_has_missing_formula_cache(node):
            add_factor("output_formula_cache_missing", f"{workbook_name} output formula cache is missing at {node_id}.", "weak", "value_delta")

    for node_id in _path_node_ids(paths):
        node = candidate_graph["nodes"].get(node_id) or baseline_graph["nodes"].get(node_id)
        if not node:
            continue
        formula = node.get("new_formula") or node.get("old_formula")
        if not formula:
            continue
        if formula.get("parse_status") and formula.get("parse_status") != "ok":
            add_factor("partial_formula_parse", f"Formula dependency extraction was partial at {node_id}.", "weak", "dependency")
        if formula.get("has_dynamic_reference"):
            add_factor("dynamic_reference", f"Formula path includes dynamic reference logic at {node_id}.", "weak", "dependency")
        if formula.get("has_external_reference"):
            add_factor("external_reference", f"Formula path includes an external workbook reference at {node_id}.", "moderate", "dependency")
        if formula.get("has_volatile_function"):
            add_factor("volatile_function", f"Formula path includes a volatile function at {node_id}.", "moderate", "dependency")
        if _node_has_missing_formula_cache(node):
            add_factor("path_formula_cache_missing", f"Formula cache is missing on the path at {node_id}.", "weak", "value_delta")

    capped = strength
    for factor in factors:
        capped = _cap_strength(capped, factor["strength_cap"])
    factors = _dedupe_confidence_factors(factors)
    dependency_confidence = _path_dependency_confidence(paths)
    value_delta_confidence = "high" if _delta_is_nonzero(output_change.get("delta")) else "moderate"
    for factor in factors:
        area = factor.get("confidence_area")
        if area in {"dependency", "explanation"}:
            dependency_confidence = _cap_confidence(dependency_confidence, factor["strength_cap"])
        if area in {"value_delta", "explanation"}:
            value_delta_confidence = _cap_confidence(value_delta_confidence, factor["strength_cap"])
    return capped, factors, dependency_confidence, value_delta_confidence


def _path_node_ids(paths: Sequence[Dict[str, Any]]) -> Set[str]:
    node_ids: Set[str] = set()
    for path in paths:
        node_ids.update(path.get("nodes", []))
    return node_ids


def _cap_strength(strength: str, cap: str) -> str:
    order = {"unexplained": 0, "weak": 1, "moderate": 2, "strong": 3}
    if order.get(strength, 0) <= order.get(cap, 0):
        return strength
    return cap


def _dedupe_confidence_factors(factors: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen = set()
    result = []
    for factor in factors:
        key = (factor.get("code"), factor.get("message"))
        if key in seen:
            continue
        seen.add(key)
        result.append(factor)
    return result


def _path_dependency_confidence(paths: Sequence[Dict[str, Any]]) -> str:
    if not paths:
        return "weak"
    confidence = min(float(path.get("confidence", 0)) for path in paths)
    if confidence >= 0.85:
        return "high"
    if confidence >= 0.65:
        return "moderate"
    return "weak"


def _cap_confidence(confidence: str, cap: str) -> str:
    order = {"weak": 1, "moderate": 2, "high": 3, "strong": 3, "unexplained": 0}
    normalized_cap = "high" if cap == "strong" else cap
    if order.get(confidence, 0) <= order.get(normalized_cap, 0):
        return confidence
    return normalized_cap


def _node_has_missing_formula_cache(node: Optional[Dict[str, Any]]) -> bool:
    if not node:
        return False
    if not (node.get("new_formula") or node.get("old_formula")):
        return False
    value = node.get("new_value") or node.get("old_value") or {}
    return value.get("value_type") == "blank" or (value.get("typed_value") is None and not value.get("display_value"))


def build_change_impact_dag(
    changes: Sequence[Dict[str, Any]],
    impacted_outputs: Sequence[Dict[str, Any]],
    candidate_graph: Dict[str, Any],
    baseline_graph: Dict[str, Any],
) -> Dict[str, Any]:
    included_nodes: Set[str] = set()
    included_edges: Set[str] = set()
    roots = _ordered_graph_roots(changes)[:40]
    outputs = [output["ref"] for output in impacted_outputs[:40]]
    for output in impacted_outputs[:40]:
        for path in output.get("representative_paths", []):
            included_nodes.update(path["nodes"])
            included_edges.update(path["edges"])
    included_nodes.update(roots)
    included_nodes.update(outputs)

    graph_edges_by_id = {edge["id"]: edge for edge in candidate_graph["edges"]}
    graph_edges_by_id.update({edge["id"]: edge for edge in baseline_graph["edges"] if edge["id"] not in graph_edges_by_id})
    nodes = {}
    for node_id in included_nodes:
        node = candidate_graph["nodes"].get(node_id) or baseline_graph["nodes"].get(node_id) or _placeholder_node_dict(node_id)
        node = dict(node)
        node["changes"] = [change["id"] for change in changes if node_id in _change_graph_root_refs(change) or change["object_ref"] == node_id]
        nodes[node_id] = node
    edges = [graph_edges_by_id[edge_id] for edge_id in sorted(included_edges) if edge_id in graph_edges_by_id]
    compaction = _compact_impact_paths(impacted_outputs[:40], nodes)
    return {
        "roots": [root for root in roots if root in nodes],
        "outputs": [output for output in outputs if output in nodes],
        "nodes": nodes,
        "edges": edges,
        "compacted_nodes": compaction["nodes"],
        "compacted_edges": compaction["edges"],
        "compacted_paths": compaction["paths"],
        "compaction_groups": compaction["groups"],
        "range_membership_summary": candidate_graph.get("range_membership_summary") or baseline_graph.get("range_membership_summary") or [],
        "collapsed_node_count": compaction["collapsed_node_count"],
        "omitted_node_count": max(0, len(candidate_graph["nodes"]) - len(nodes)),
    }


def _compact_impact_paths(impacted_outputs: Sequence[Dict[str, Any]], nodes: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    group_by_members: Dict[Tuple[str, ...], str] = {}
    groups: Dict[str, Dict[str, Any]] = {}
    compact_nodes: Dict[str, Dict[str, Any]] = {}
    compact_edges_seen: Set[Tuple[str, str]] = set()
    compact_edges: List[Dict[str, Any]] = []
    compact_paths: List[Dict[str, Any]] = []

    def add_node(node_id: str) -> None:
        if node_id in compact_nodes:
            return
        if node_id in groups:
            compact_nodes[node_id] = groups[node_id]
        else:
            compact_nodes[node_id] = nodes.get(node_id, _placeholder_node_dict(node_id))

    def flush_buffer(buffer: List[str], compacted: List[str]) -> None:
        if not buffer:
            return
        if len(buffer) == 1:
            compacted.append(buffer[0])
            return
        key = tuple(buffer)
        group_id = group_by_members.get(key)
        if group_id is None:
            group_id = f"collapsed:{len(group_by_members) + 1:03d}"
            group_by_members[key] = group_id
            groups[group_id] = _compaction_group(group_id, buffer, nodes)
        compacted.append(group_id)

    for output in impacted_outputs:
        for path in output.get("representative_paths", []):
            raw_nodes = list(path.get("nodes", []))
            if not raw_nodes:
                continue
            compacted: List[str] = []
            buffer: List[str] = []
            for index, node_id in enumerate(raw_nodes):
                terminal = index == 0 or index == len(raw_nodes) - 1
                if not terminal and _node_is_compactable(node_id, nodes.get(node_id, {})):
                    buffer.append(node_id)
                    continue
                flush_buffer(buffer, compacted)
                buffer = []
                compacted.append(node_id)
            flush_buffer(buffer, compacted)
            compacted = _dedupe_consecutive(compacted)
            for node_id in compacted:
                add_node(node_id)
            for left, right in zip(compacted, compacted[1:]):
                if left == right or (left, right) in compact_edges_seen:
                    continue
                compact_edges_seen.add((left, right))
                compact_edges.append({"from": left, "to": right, "edge_type": "compacted_dependency"})
            compact_paths.append(
                {
                    "output_ref": output.get("ref"),
                    "path_id": path.get("id"),
                    "nodes": compacted,
                    "raw_node_count": len(raw_nodes),
                    "collapsed_node_count": max(0, len(raw_nodes) - len(compacted)),
                }
            )

    collapsed_count = sum(len(group.get("member_nodes", [])) for group in groups.values())
    return {
        "nodes": compact_nodes,
        "edges": compact_edges,
        "paths": compact_paths,
        "groups": list(groups.values()),
        "collapsed_node_count": collapsed_count,
    }


def _node_is_compactable(node_id: str, node: Dict[str, Any]) -> bool:
    node_type = node.get("node_type")
    if node_type in {"defined_name", "table", "range", "external_reference", "table_column"}:
        return False
    if node.get("semantic_role") in {"assumption", "raw_data", "output"}:
        return False
    return "!" in node_id


def _compaction_group(group_id: str, member_nodes: Sequence[str], nodes: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    member_list = list(member_nodes)
    range_ref = _bounding_range(member_list)
    first_node = nodes.get(member_list[0], {}) if member_list else {}
    sheet = split_cell_id(member_list[0])[0] if member_list else ""
    has_formula = any((nodes.get(node_id, {}).get("new_formula") or nodes.get(node_id, {}).get("old_formula")) for node_id in member_list)
    group_type = "formula_block" if has_formula else "intermediate_chain"
    label_prefix = "Formula block" if has_formula else "Intermediate chain"
    return {
        "id": group_id,
        "ref": range_ref,
        "node_type": "collapsed_block",
        "group_type": group_type,
        "label": f"{label_prefix}: {range_ref}",
        "semantic_role": first_node.get("semantic_role", "intermediate_calculation"),
        "sheet_name": sheet or first_node.get("sheet_name"),
        "member_count": len(member_list),
        "member_nodes": member_list,
        "range_ref": range_ref,
        "changes": [],
        "materiality_score": max((nodes.get(node_id, {}).get("materiality_score", 0) for node_id in member_list), default=0),
        "confidence": min((nodes.get(node_id, {}).get("confidence", 1.0) for node_id in member_list), default=1.0),
    }


def _dedupe_consecutive(items: Sequence[str]) -> List[str]:
    result: List[str] = []
    for item in items:
        if result and result[-1] == item:
            continue
        result.append(item)
    return result


def build_summary(
    changes: Sequence[Dict[str, Any]],
    impacted_outputs: Sequence[Dict[str, Any]],
    baseline: Any,
    candidate: Any,
    diagnostics: Sequence[Dict[str, Any]],
    alignment: Dict[str, Any],
) -> Dict[str, Any]:
    direct_count = sum(1 for change in changes if change["directness"] == "direct")
    propagated_count = sum(1 for change in changes if change["directness"] == "propagated")
    unexplained_count = sum(1 for change in changes if change["directness"] == "unexplained")
    warning_count = sum(1 for diagnostic in diagnostics if diagnostic.get("severity") == "warning")
    confidence = "high"
    if warning_count or unexplained_count:
        confidence = "medium"
    if any(diagnostic.get("severity") == "error" for diagnostic in diagnostics) or warning_count > 20:
        confidence = "low"
    one_sentence = (
        f"Detected {direct_count} direct changes, {propagated_count} propagated value changes, "
        f"and {unexplained_count} unexplained value changes."
    )
    alignment_summary = alignment.get("summary", {})
    return {
        "direct_change_count": direct_count,
        "propagated_change_count": propagated_count,
        "unexplained_change_count": unexplained_count,
        "structural_matches": alignment_summary.get("matched_cells", 0),
        "semantic_matches": alignment_summary.get("semantic_matches", 0),
        "shifted_semantic_matches": alignment_summary.get("shifted_semantic_matches", 0),
        "unmatched_old_cells": alignment_summary.get("unmatched_old_cells", 0),
        "unmatched_new_cells": alignment_summary.get("unmatched_new_cells", 0),
        "sheets_added": sum(1 for change in changes if change["kind"] == "sheet_added"),
        "sheets_deleted": sum(1 for change in changes if change["kind"] == "sheet_deleted"),
        "formulas_changed": sum(1 for change in changes if change["kind"].startswith("formula_") or change["kind"] == "formula_changed"),
        "constants_changed": sum(1 for change in changes if change["kind"] == "constant_changed"),
        "cached_values_changed": sum(1 for change in changes if change["kind"] == "cached_value_changed"),
        "outputs_changed": len(impacted_outputs),
        "has_macros": bool(baseline.file.get("has_macros") or candidate.file.get("has_macros")),
        "has_external_links": any(diagnostic.get("code") == "EXTERNAL_LINKS_DETECTED" for diagnostic in diagnostics),
        "has_opaque_formulas": any(diagnostic.get("code") == "OPAQUE_DYNAMIC_REFERENCE" for diagnostic in diagnostics),
        "confidence": confidence,
        "one_sentence_summary": one_sentence,
    }


def build_llm_summary(
    changes: Sequence[Dict[str, Any]],
    impacted_outputs: Sequence[Dict[str, Any]],
    grouped_changes: Sequence[Dict[str, Any]],
    summary: Dict[str, Any],
    diagnostics: Sequence[Dict[str, Any]],
    config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    suppressed_ids = _suppressed_change_ids_from_groups(grouped_changes)
    direct_changes = [change for change in changes if change.get("directness") == "direct" and change.get("id") not in suppressed_ids]
    top_direct = direct_changes[:8]
    top_groups = list(grouped_changes[:8])
    final_outputs, impacted_intermediates = split_impacted_outputs(impacted_outputs, config or {})
    top_outputs = list(final_outputs[:8])
    top_intermediates = list(impacted_intermediates[:8])
    warning_codes = sorted({diagnostic.get("code", "") for diagnostic in diagnostics if diagnostic.get("severity") == "warning" and diagnostic.get("code")})
    caveats = [
        "Uses cached workbook formula values; numeric deltas assume both workbooks were saved after recalculation.",
    ]
    if summary.get("unexplained_change_count"):
        caveats.append(f"{summary['unexplained_change_count']} changed value(s) had no detected upstream explanation.")
    if warning_codes:
        caveats.append("Warnings present: " + ", ".join(warning_codes) + ".")
    output_factor_codes = sorted(
        {
            factor.get("code", "")
            for output in top_outputs
            for factor in output.get("confidence_factors", [])
            if factor.get("code")
        }
    )
    if output_factor_codes:
        caveats.append("Output confidence factors: " + ", ".join(output_factor_codes) + ".")

    one_sentence = _llm_one_sentence(summary, top_direct, top_outputs, top_groups)
    return {
        "schema_version": "0.1",
        "intended_use": "A compact, deterministic context object for LLMs to answer a user with one sentence plus optional evidence.",
        "one_sentence_summary": one_sentence,
        "confidence": summary.get("confidence", "unknown"),
        "counts": {
            "direct_changes": len(direct_changes),
            "raw_direct_changes": summary.get("direct_change_count", 0),
            "propagated_changes": summary.get("propagated_change_count", 0),
            "unexplained_changes": summary.get("unexplained_change_count", 0),
            "formula_changes": summary.get("formulas_changed", 0),
            "outputs_changed": summary.get("outputs_changed", 0),
            "final_outputs_changed": len(final_outputs),
            "impacted_intermediates": len(impacted_intermediates),
            "shifted_semantic_matches": summary.get("shifted_semantic_matches", 0),
        },
        "top_direct_changes": [_llm_change_fact(change) for change in top_direct],
        "top_change_groups": [_llm_group_fact(group) for group in top_groups],
        "final_outputs": [_llm_output_fact(output) for output in top_outputs],
        "impacted_intermediates": [_llm_output_fact(output) for output in top_intermediates],
        "top_impacted_outputs": [_llm_output_fact(output) for output in top_outputs],
        "claims": _llm_claims(top_direct, top_groups, top_outputs),
        "caveats": caveats,
        "safe_response_rules": [
            "Use the one_sentence_summary verbatim or paraphrase it without adding unsupported causality.",
            "Say 'likely explains' only when explanation_strength is strong.",
            "Say 'associated with' for moderate or weak paths.",
            "Always include cell references for material claims.",
        ],
    }


def _llm_one_sentence(
    summary: Dict[str, Any],
    top_direct: Sequence[Dict[str, Any]],
    top_outputs: Sequence[Dict[str, Any]],
    top_groups: Sequence[Dict[str, Any]] = (),
) -> str:
    if summary.get("direct_change_count", 0) == 0 and summary.get("propagated_change_count", 0) == 0 and summary.get("unexplained_change_count", 0) == 0:
        return "No workbook changes were detected."
    if top_outputs:
        output = top_outputs[0]
        output_ref = output.get("ref", "an output")
        output_label = output.get("label") or output_ref
        old_display = (output.get("old_value") or {}).get("display_value", "")
        new_display = (output.get("new_value") or {}).get("display_value", "")
        delta = _delta_text(output)
        strength = output.get("explanation_strength", "unknown")
        named_rebound = next((group for group in top_groups if group.get("group_type") == "named_range_rebound"), None)
        direct_phrase = _named_range_rebound_phrase(named_rebound) if named_rebound else _direct_change_phrase(top_direct)
        explanation = "likely explained by" if strength == "strong" else "associated with"
        if strength == "unexplained":
            explanation = "without a detected upstream explanation despite"
        if not delta and old_display == new_display and output.get("change_kind", "").startswith("formula"):
            return (
                f"{output_ref} {output_label} formula changed while the cached value stayed at {old_display or 'blank'}, "
                f"{explanation} {direct_phrase}; {summary.get('unexplained_change_count', 0)} unexplained value changes were detected."
            )
        return (
            f"{output_ref} {output_label} changed from {old_display or 'blank'} to {new_display or 'blank'}"
            f"{f' ({delta})' if delta else ''}, {explanation} {direct_phrase}; "
            f"{summary.get('unexplained_change_count', 0)} unexplained value changes were detected."
        )
    if top_direct:
        noun = "change" if summary.get("direct_change_count", 0) == 1 else "changes"
        return (
            f"Detected {summary.get('direct_change_count', 0)} direct {noun}, led by {_change_short(top_direct[0])}, "
            f"with {summary.get('propagated_change_count', 0)} propagated value changes and "
            f"{summary.get('unexplained_change_count', 0)} unexplained value changes."
        )
    return (
        f"Detected {summary.get('direct_change_count', 0)} direct changes, "
        f"{summary.get('propagated_change_count', 0)} propagated value changes, and "
        f"{summary.get('unexplained_change_count', 0)} unexplained value changes."
    )


def _output_path_length(output: Dict[str, Any]) -> int:
    paths = output.get("representative_paths") or []
    if not paths:
        return 0
    return max(len(path.get("nodes", [])) for path in paths)


def _configured_output_refs(config: Dict[str, Any]) -> Set[str]:
    nested = config.get("workbook_diff", config)
    refs: Set[str] = set()
    for output in nested.get("outputs", []) or []:
        if isinstance(output, dict) and output.get("ref"):
            refs.add(output["ref"])
        elif isinstance(output, str):
            refs.add(output)
    return refs


def _configured_output_names(config: Dict[str, Any]) -> Dict[str, str]:
    nested = config.get("workbook_diff", config)
    names: Dict[str, str] = {}
    for output in nested.get("outputs", []) or []:
        if not isinstance(output, dict):
            continue
        ref = output.get("ref")
        name = output.get("name") or output.get("label")
        if ref and name:
            names[ref] = str(name)
    return names


def _rank_outputs_for_summary(impacted_outputs: Sequence[Dict[str, Any]], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    configured_output_refs = _configured_output_refs(config)
    return sorted(
        impacted_outputs,
        key=lambda output: (
            0 if output.get("ref") in configured_output_refs else 1,
            _summary_output_priority(output.get("ref", "")),
            0 if output.get("semantic_role") == "output" else 1,
            -_output_path_length(output),
            -float(output.get("materiality_score") or 0),
            output.get("ref", ""),
        ),
    )


def _path_intermediate_output_refs(outputs: Sequence[Dict[str, Any]]) -> Set[str]:
    output_refs = {output.get("ref") for output in outputs if output.get("ref")}
    intermediate_refs: Set[str] = set()
    for output in outputs:
        for path in output.get("representative_paths", []) or []:
            nodes = path.get("nodes") or []
            for node_id in nodes[:-1]:
                if node_id in output_refs:
                    intermediate_refs.add(node_id)
    return intermediate_refs


def _is_final_output(output: Dict[str, Any], configured_refs: Set[str], intermediate_refs: Set[str]) -> bool:
    ref = output.get("ref", "")
    if ref in configured_refs:
        return True
    if ref in intermediate_refs:
        return False
    sheet, _address = split_cell_id(ref) if "!" in ref else ("", "")
    if sheet.upper() in {"SUMMARY", "DASHBOARD", "OUTPUT", "BOARD", "KPIS", "REPORT"}:
        return True
    return output.get("semantic_role") == "output"


def _summary_output_priority(ref: str) -> int:
    sheet, address = split_cell_id(ref)
    digits = "".join(ch for ch in address if ch.isdigit())
    row = int(digits) if digits else 10_000
    if sheet.upper() in {"SUMMARY", "DASHBOARD", "OUTPUT", "BOARD", "KPIS", "REPORT"} and row <= 15:
        return 0
    return 1


def _llm_change_fact(change: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": change.get("id"),
        "kind": change.get("kind"),
        "ref": change.get("object_ref"),
        "old_ref": change.get("old_ref"),
        "new_ref": change.get("new_ref"),
        "address_changed": change.get("address_changed"),
        "semantic_id": change.get("semantic_id"),
        "label": _best_label_from_change(change),
        "old": change.get("old_display"),
        "new": change.get("new_display"),
        "delta": _delta_text(change),
        "semantic_role": change.get("semantic_role"),
        "materiality_score": change.get("materiality_score"),
        "confidence": change.get("confidence"),
    }


def _llm_group_fact(group: Dict[str, Any]) -> Dict[str, Any]:
    fact = {
        "id": group.get("id"),
        "group_type": group.get("group_type"),
        "label": group.get("label"),
        "range_ref": group.get("range_ref"),
        "change_count": group.get("change_count"),
        "summary": group.get("summary"),
        "materiality_score": group.get("materiality_score"),
    }
    for key in ["primary_change_id", "name", "old_target", "new_target", "effective_old", "effective_new", "delta"]:
        if key in group:
            fact[key] = group.get(key)
    return fact


def _llm_output_fact(output: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "ref": output.get("ref"),
        "old_ref": output.get("old_ref"),
        "new_ref": output.get("new_ref"),
        "address_changed": output.get("address_changed"),
        "change_id": output.get("change_id"),
        "change_kind": output.get("change_kind"),
        "semantic_id": output.get("semantic_id"),
        "label": output.get("label"),
        "old": (output.get("old_value") or {}).get("display_value"),
        "new": (output.get("new_value") or {}).get("display_value"),
        "delta": _delta_text(output),
        "explanation_strength": output.get("explanation_strength"),
        "dependency_confidence": output.get("dependency_confidence"),
        "value_delta_confidence": output.get("value_delta_confidence"),
        "confidence_factors": output.get("confidence_factors", []),
        "upstream_change_ids": output.get("upstream_change_ids", []),
        "representative_paths": output.get("representative_paths", [])[:3],
        "explanation": output.get("explanation"),
    }


def _direct_change_phrase(top_direct: Sequence[Dict[str, Any]]) -> str:
    if not top_direct:
        return "the detected direct changes"
    if len(top_direct) == 1:
        return _change_short(top_direct[0])
    return f"{_change_short(top_direct[0])} and {len(top_direct) - 1} other top direct change(s)"


def _change_short(change: Dict[str, Any]) -> str:
    ref = change.get("object_ref", "a changed cell")
    if change.get("kind") == "defined_name_changed":
        return f"named range {ref} rebinding from {change.get('old_display') or 'blank'} to {change.get('new_display') or 'blank'}"
    label = _best_label_from_change(change)
    old_display = change.get("old_display", "")
    new_display = change.get("new_display", "")
    if label and label != ref:
        return f"{ref} {label} changing from {old_display or 'blank'} to {new_display or 'blank'}"
    return f"{ref} changing from {old_display or 'blank'} to {new_display or 'blank'}"


def _named_range_rebound_phrase(group: Optional[Dict[str, Any]]) -> str:
    if not group:
        return ""
    phrase = f"named range {group.get('name')} rebinding from {group.get('old_target')} to {group.get('new_target')}"
    old_display = group.get("effective_old")
    new_display = group.get("effective_new")
    if old_display or new_display:
        phrase += f", moving the effective value from {old_display or 'blank'} to {new_display or 'blank'}"
    return phrase


def _llm_claims(
    top_direct: Sequence[Dict[str, Any]],
    top_groups: Sequence[Dict[str, Any]],
    top_outputs: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    claims: List[Dict[str, Any]] = []
    for group in top_groups:
        if group.get("group_type") != "named_range_rebound":
            continue
        claims.append(
            {
                "claim": _named_range_rebound_summary(
                    str(group.get("name", "")),
                    str(group.get("old_target", "")),
                    str(group.get("new_target", "")),
                    str(group.get("effective_old") or ""),
                    str(group.get("effective_new") or ""),
                    group.get("delta"),
                ),
                "refs": [ref for ref in [group.get("name"), group.get("old_target"), group.get("new_target")] if ref],
                "change_ids": [group.get("primary_change_id")] if group.get("primary_change_id") else [],
                "evidence_type": "named_range_rebinding",
                "confidence": "moderate",
                "caveats": ["non_cell_root"],
            }
        )
    for output in top_outputs:
        refs = [output.get("ref")] if output.get("ref") else []
        old_value = output.get("old_value") or {}
        new_value = output.get("new_value") or {}
        claims.append(
            {
                "claim": f"{output.get('label') or output.get('ref')} changed from {old_value.get('display_value') or 'blank'} to {new_value.get('display_value') or 'blank'}",
                "refs": refs,
                "old_value": old_value.get("typed_value"),
                "new_value": new_value.get("typed_value"),
                "delta": output.get("delta"),
                "evidence_type": "cached_value_diff",
                "confidence": output.get("value_delta_confidence") or output.get("explanation_strength"),
                "caveats": ["cached_value_mode"] + [factor.get("code") for factor in output.get("confidence_factors", []) if factor.get("code")],
            }
        )
        if output.get("upstream_change_ids") and output.get("explanation_strength") != "unexplained":
            claims.append(
                {
                    "claim": f"Detected upstream changes are associated with {output.get('label') or output.get('ref')}",
                    "refs": refs,
                    "root_change_ids": output.get("upstream_change_ids", []),
                    "representative_paths": output.get("representative_paths", [])[:3],
                    "evidence_type": "dependency_path",
                    "confidence": output.get("explanation_strength"),
                    "dependency_confidence": output.get("dependency_confidence"),
                    "value_delta_confidence": output.get("value_delta_confidence"),
                    "caveats": [factor.get("code") for factor in output.get("confidence_factors", []) if factor.get("code")],
                }
            )
    return claims[:12]


def _delta_text(item: Dict[str, Any]) -> str:
    delta = item.get("delta")
    if not delta:
        return ""
    if not _delta_is_nonzero(delta):
        return ""
    parts = [delta.get("display_point_delta") or delta.get("display_delta"), delta.get("display_relative_delta")]
    return " / ".join(part for part in parts if part)


def _delta_is_nonzero(delta: Optional[Dict[str, Any]]) -> bool:
    if not delta:
        return False
    absolute_delta = delta.get("absolute_delta")
    return isinstance(absolute_delta, (int, float)) and abs(float(absolute_delta)) > 1e-12


def _diff_sheets(baseline: Any, candidate: Any, changes: List[Dict[str, Any]], counter: int) -> int:
    old_by_name = {sheet["name"]: sheet for sheet in baseline.sheets}
    new_by_name = {sheet["name"]: sheet for sheet in candidate.sheets}
    for name in sorted(set(old_by_name) - set(new_by_name)):
        changes.append(_workbook_change(counter, "sheet_deleted", name, old_by_name[name], None))
        counter += 1
    for name in sorted(set(new_by_name) - set(old_by_name)):
        changes.append(_workbook_change(counter, "sheet_added", name, None, new_by_name[name]))
        counter += 1
    for name in sorted(set(old_by_name) & set(new_by_name)):
        if old_by_name[name]["visibility"] != new_by_name[name]["visibility"]:
            changes.append(_workbook_change(counter, "sheet_visibility_changed", name, old_by_name[name]["visibility"], new_by_name[name]["visibility"]))
            counter += 1
    return counter


def _diff_defined_names(baseline: Any, candidate: Any, changes: List[Dict[str, Any]], counter: int) -> int:
    old = {item["name"].upper(): item for item in baseline.defined_names}
    new = {item["name"].upper(): item for item in candidate.defined_names}
    for name in sorted(set(old) - set(new)):
        changes.append(_name_change(counter, "defined_name_deleted", old[name]["name"], old[name], None))
        counter += 1
    for name in sorted(set(new) - set(old)):
        changes.append(_name_change(counter, "defined_name_added", new[name]["name"], None, new[name]))
        counter += 1
    for name in sorted(set(old) & set(new)):
        if old[name].get("ref") != new[name].get("ref"):
            changes.append(_name_change(counter, "defined_name_changed", new[name]["name"], old[name].get("ref"), new[name].get("ref")))
            counter += 1
    return counter


def _diff_tables(baseline: Any, candidate: Any, changes: List[Dict[str, Any]], counter: int) -> int:
    old = {item["name"].upper(): item for item in baseline.tables}
    new = {item["name"].upper(): item for item in candidate.tables}
    for name in sorted(set(old) - set(new)):
        changes.append(_table_change(counter, "table_deleted", old[name]["name"], old[name], None))
        counter += 1
    for name in sorted(set(new) - set(old)):
        changes.append(_table_change(counter, "table_added", new[name]["name"], None, new[name]))
        counter += 1
    for name in sorted(set(old) & set(new)):
        if old[name].get("ref") != new[name].get("ref"):
            changes.append(_table_change(counter, "table_range_changed", new[name]["name"], old[name].get("ref"), new[name].get("ref")))
            counter += 1
    return counter


def _diff_workbook_file_info(baseline: Any, candidate: Any, changes: List[Dict[str, Any]], counter: int) -> int:
    if bool(baseline.file.get("has_macros")) != bool(candidate.file.get("has_macros")):
        changes.append(_workbook_change(counter, "macro_presence_changed", "workbook", baseline.file.get("has_macros"), candidate.file.get("has_macros")))
        counter += 1
    if baseline.file.get("calc_mode") != candidate.file.get("calc_mode"):
        changes.append(_workbook_change(counter, "calculation_mode_changed", "workbook", baseline.file.get("calc_mode"), candidate.file.get("calc_mode")))
        counter += 1
    return counter


def _diff_cells(
    baseline: Any,
    candidate: Any,
    alignment: Dict[str, Any],
    changes: List[Dict[str, Any]],
    counter: int,
    options: Dict[str, Any],
) -> int:
    old_sheet_names = [sheet["name"] for sheet in baseline.sheets]
    new_sheet_names = [sheet["name"] for sheet in candidate.sheets]
    for match in alignment["matches"]:
        old_cell = baseline.cells.get(match["old_ref"])
        new_cell = candidate.cells.get(match["new_ref"])
        if old_cell is None or new_cell is None:
            continue
        counter = _diff_matched_cell(old_cell, new_cell, old_sheet_names, new_sheet_names, alignment, changes, counter, options)

    for ref in alignment["unmatched_new_refs"]:
        new_cell = candidate.cells.get(ref)
        if new_cell is None:
            continue
        changes.append(_cell_change(counter, "cell_added", ref, None, new_cell.value.to_dict(), new_cell, None, new_cell, directness="direct"))
        counter += 1

    for ref in alignment["unmatched_old_refs"]:
        old_cell = baseline.cells.get(ref)
        if old_cell is None:
            continue
        changes.append(_cell_change(counter, "cell_deleted", ref, old_cell.value.to_dict(), None, old_cell, old_cell, None, directness="direct"))
        counter += 1
    return counter


def _diff_matched_cell(
    old_cell: Any,
    new_cell: Any,
    old_sheet_names: Sequence[str],
    new_sheet_names: Sequence[str],
    alignment: Dict[str, Any],
    changes: List[Dict[str, Any]],
    counter: int,
    options: Dict[str, Any],
) -> int:
    ref = new_cell.id
    if old_cell.kind != new_cell.kind:
        if old_cell.kind == "formula" and new_cell.kind == "constant":
            kind = "formula_to_constant"
        elif old_cell.kind == "constant" and new_cell.kind == "formula":
            kind = "constant_to_formula"
        else:
            kind = "cell_changed"
        changes.append(
            _cell_change(
                counter,
                kind,
                ref,
                _cell_payload(old_cell),
                _cell_payload(new_cell),
                new_cell,
                old_cell,
                new_cell,
                directness="direct",
            )
        )
        counter += 1
        return counter

    if old_cell.kind == "formula" and new_cell.kind == "formula":
        old_formula = old_cell.formula.raw if old_cell.formula else None
        new_formula = new_cell.formula.raw if new_cell.formula else None
        if normalize_formula(old_formula) != normalize_formula(new_formula):
            if not _formula_equal_after_alignment(old_formula, new_formula, old_cell, new_cell, old_sheet_names, new_sheet_names, alignment):
                formula_diff = classify_formula_change(old_formula, new_formula, new_cell.sheet_name, new_sheet_names)
                changes.append(
                    _cell_change(
                        counter,
                        formula_diff["kind"],
                        ref,
                        old_formula,
                        new_formula,
                        new_cell,
                        old_cell,
                        new_cell,
                        directness="direct",
                        evidence=[{"type": "formula_diff", **formula_diff}],
                    )
                )
                counter += 1
        elif _value_changed(old_cell.value.to_dict(), new_cell.value.to_dict()):
            changes.append(
                _cell_change(
                    counter,
                    "cached_value_changed",
                    ref,
                    old_cell.value.to_dict(),
                    new_cell.value.to_dict(),
                    new_cell,
                    old_cell,
                    new_cell,
                    directness="unexplained",
                )
            )
            counter += 1
    elif old_cell.kind == "constant" and new_cell.kind == "constant":
        if _value_changed(old_cell.value.to_dict(), new_cell.value.to_dict()):
            kind = "error_changed" if old_cell.value.value_type == "error" or new_cell.value.value_type == "error" else "constant_changed"
            changes.append(
                _cell_change(
                    counter,
                    kind,
                    ref,
                    old_cell.value.to_dict(),
                    new_cell.value.to_dict(),
                    new_cell,
                    old_cell,
                    new_cell,
                    directness="direct",
                )
            )
            counter += 1

    if options.get("include_style_changes") and old_cell.style != new_cell.style:
        changes.append(_cell_change(counter, "style_changed", ref, old_cell.style, new_cell.style, new_cell, old_cell, new_cell, directness="direct"))
        counter += 1
    if options.get("include_comments") and old_cell.comment != new_cell.comment:
        changes.append(_cell_change(counter, "comment_changed", ref, old_cell.comment, new_cell.comment, new_cell, old_cell, new_cell, directness="direct"))
        counter += 1
    if old_cell.hyperlink != new_cell.hyperlink:
        changes.append(_cell_change(counter, "hyperlink_changed", ref, old_cell.hyperlink, new_cell.hyperlink, new_cell, old_cell, new_cell, directness="direct"))
        counter += 1
    return counter


def _formula_equal_after_alignment(
    old_formula: Optional[str],
    new_formula: Optional[str],
    old_cell: Any,
    new_cell: Any,
    old_sheet_names: Sequence[str],
    new_sheet_names: Sequence[str],
    alignment: Dict[str, Any],
) -> bool:
    old_refs, _, _, _ = parse_formula_references(old_formula, old_cell.sheet_name, old_sheet_names)
    new_refs, _, _, _ = parse_formula_references(new_formula, new_cell.sheet_name, new_sheet_names)
    remapped_old_targets = sorted(_remap_formula_target(ref.target, alignment["old_to_new"]) for ref in old_refs)
    new_targets = sorted(ref.target for ref in new_refs)
    if remapped_old_targets != new_targets:
        return False
    return (
        formula_functions(old_formula or "") == formula_functions(new_formula or "")
        and numeric_constants(old_formula or "") == numeric_constants(new_formula or "")
        and formula_operators(old_formula or "") == formula_operators(new_formula or "")
    )


def _remap_formula_target(target: str, old_to_new: Dict[str, str]) -> str:
    if target in old_to_new:
        return old_to_new[target]
    if "!" not in target or ":" not in target:
        return target
    sheet, cells = split_cell_id(target)
    start, end = cells.split(":", 1)
    mapped_start = old_to_new.get(f"{sheet}!{start}")
    mapped_end = old_to_new.get(f"{sheet}!{end}")
    if not mapped_start or not mapped_end:
        return target
    mapped_start_sheet, mapped_start_address = split_cell_id(mapped_start)
    mapped_end_sheet, mapped_end_address = split_cell_id(mapped_end)
    if mapped_start_sheet != mapped_end_sheet:
        return target
    return f"{mapped_start_sheet}!{mapped_start_address}:{mapped_end_address}"


def _cell_change(
    counter: int,
    kind: str,
    ref: str,
    old: Any,
    new: Any,
    display_cell: Any,
    old_cell: Any,
    new_cell: Any,
    directness: str,
    evidence: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    old_value = old_cell.value.to_dict() if old_cell else None
    new_value = new_cell.value.to_dict() if new_cell else None
    delta = numeric_delta(old_value, new_value)
    evidence = list(evidence or [])
    display_identity = (display_cell.semantic_identity if display_cell else {}) or {}
    semantic_id = (display_cell.semantic_id if display_cell else None) or (old_cell.semantic_id if old_cell else None)
    old_ref = old_cell.id if old_cell else None
    new_ref = new_cell.id if new_cell else None
    semantic_role = display_cell.semantic_role if display_cell else "unknown"
    if semantic_role == "unknown" and old_cell is not None:
        semantic_role = old_cell.semantic_role
    if old_ref and new_ref and old_ref != new_ref:
        evidence.insert(
            0,
            {
                "type": "structural_alignment",
                "old_ref": old_ref,
                "new_ref": new_ref,
                "semantic_id": semantic_id,
                "confidence": display_identity.get("confidence"),
            },
        )
    return {
        "id": f"chg_{counter:04d}",
        "kind": kind,
        "scope": "cell",
        "object_ref": ref,
        "old_ref": old_ref,
        "new_ref": new_ref,
        "address_changed": bool(old_ref and new_ref and old_ref != new_ref),
        "semantic_id": semantic_id,
        "semantic_identity": display_identity,
        "sheet_name": display_cell.sheet_name if display_cell else split_cell_id(ref)[0],
        "range_ref": ref,
        "old": old,
        "new": new,
        "old_display": _display_from_value(old),
        "new_display": _display_from_value(new),
        "delta": delta,
        "labels": display_cell.labels if display_cell else [],
        "semantic_role": semantic_role,
        "directness": directness,
        "materiality_score": 0,
        "confidence": 0.95 if directness == "direct" else 0.7,
        "evidence": evidence,
        "warnings": [],
        "visibility_context": display_cell.visibility_context if display_cell else {},
    }


def _workbook_change(counter: int, kind: str, ref: str, old: Any, new: Any) -> Dict[str, Any]:
    return {
        "id": f"chg_{counter:04d}",
        "kind": kind,
        "scope": "workbook" if ref == "workbook" else "sheet",
        "object_ref": ref,
        "old": old,
        "new": new,
        "old_display": display_scalar(old),
        "new_display": display_scalar(new),
        "labels": [],
        "semantic_role": "metadata",
        "directness": "direct",
        "materiality_score": 0,
        "confidence": 0.95,
        "evidence": [],
        "warnings": [],
    }


def _name_change(counter: int, kind: str, ref: str, old: Any, new: Any) -> Dict[str, Any]:
    return {
        "id": f"chg_{counter:04d}",
        "kind": kind,
        "scope": "name",
        "object_ref": ref,
        "old": old,
        "new": new,
        "old_display": display_scalar(old),
        "new_display": display_scalar(new),
        "labels": [{"text": ref, "source": "defined_name", "confidence": 0.95}],
        "semantic_role": "unknown",
        "directness": "direct",
        "materiality_score": 0,
        "confidence": 0.9,
        "evidence": [],
        "warnings": [],
    }


def _table_change(counter: int, kind: str, ref: str, old: Any, new: Any) -> Dict[str, Any]:
    return {
        "id": f"chg_{counter:04d}",
        "kind": kind,
        "scope": "table",
        "object_ref": ref,
        "old": old,
        "new": new,
        "old_display": display_scalar(old),
        "new_display": display_scalar(new),
        "labels": [{"text": ref, "source": "table_header", "confidence": 0.85}],
        "semantic_role": "raw_data",
        "directness": "direct",
        "materiality_score": 0,
        "confidence": 0.9,
        "evidence": [],
        "warnings": [],
    }


def numeric_delta(old_value: Optional[Dict[str, Any]], new_value: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not old_value or not new_value:
        return None
    old_number = safe_float(old_value.get("typed_value"))
    new_number = safe_float(new_value.get("typed_value"))
    if old_number is None or new_number is None:
        return None
    absolute_delta = new_number - old_number
    delta = {
        "old_number": old_number,
        "new_number": new_number,
        "absolute_delta": absolute_delta,
        "display_delta": _signed_number(absolute_delta),
    }
    if old_number != 0:
        relative = new_number / old_number - 1
        delta["relative_delta"] = relative
        delta["display_relative_delta"] = _signed_percent(relative)
    number_format = (new_value.get("number_format") or old_value.get("number_format") or "").lower()
    if "%" in number_format or (abs(old_number) <= 1 and abs(new_number) <= 1):
        delta["display_point_delta"] = _signed_number(absolute_delta * 100, suffix=" pp")
    return delta


def _value_changed(old: Dict[str, Any], new: Dict[str, Any]) -> bool:
    return old.get("typed_value") != new.get("typed_value") or old.get("value_type") != new.get("value_type")


def _cell_payload(cell: Any) -> Dict[str, Any]:
    payload = {"kind": cell.kind, "value": cell.value.to_dict()}
    if cell.formula:
        payload["formula"] = cell.formula.to_dict()
    return payload


def _dependency_node(ref: str, cell: Any) -> Dict[str, Any]:
    return {
        "id": ref,
        "ref": ref,
        "node_type": "cell",
        "label": _best_label(cell.labels, ref),
        "semantic_id": cell.semantic_id,
        "semantic_identity": cell.semantic_identity,
        "semantic_role": cell.semantic_role,
        "sheet_name": cell.sheet_name,
        "old_value": None,
        "new_value": cell.value.to_dict(),
        "old_formula": None,
        "new_formula": cell.formula.to_dict() if cell.formula else None,
        "changes": [],
        "materiality_score": 0,
        "confidence": 1.0 if not cell.formula or cell.formula.parse_status == "ok" else 0.6,
    }


def _placeholder_node(ref: str) -> Dict[str, Any]:
    return _placeholder_node_dict(ref)


def _placeholder_node_dict(ref: str) -> Dict[str, Any]:
    sheet, _ = split_cell_id(ref)
    return {
        "id": ref,
        "ref": ref,
        "node_type": "cell",
        "label": ref,
        "semantic_role": "unknown",
        "sheet_name": sheet or None,
        "changes": [],
        "materiality_score": 0,
        "confidence": 0.7,
    }


def _range_or_external_node(target: str, precedent: Dict[str, Any]) -> Dict[str, Any]:
    kind = precedent.get("kind", "range")
    node_type = "range"
    if kind == "external_reference":
        node_type = "external_reference"
    elif kind == "table_reference":
        node_type = "table_column"
    elif kind == "defined_name_reference":
        node_type = "defined_name"
    return {
        "id": target,
        "ref": target,
        "node_type": node_type,
        "label": target,
        "semantic_role": "unknown",
        "changes": [],
        "materiality_score": 0,
        "confidence": precedent.get("confidence", 0.7),
    }


def _adjacency(edges: Sequence[Dict[str, Any]]) -> Dict[str, List[str]]:
    result: Dict[str, List[str]] = defaultdict(list)
    for edge in edges:
        result[edge["from"]].append(edge["to"])
    return result


def _reverse_adjacency(edges: Sequence[Dict[str, Any]]) -> Dict[str, List[str]]:
    result: Dict[str, List[str]] = defaultdict(list)
    for edge in edges:
        result[edge["to"]].append(edge["from"])
    return result


def _edge_type(kind: str) -> str:
    if kind == "cell":
        return "cell_reference"
    if kind == "range":
        return "range_reference"
    if kind == "defined_name_reference":
        return "defined_name_reference"
    if kind == "table_reference":
        return "table_reference"
    if kind == "external_reference":
        return "external_reference"
    return "dynamic_reference"


def _change_graph_root_refs(change: Dict[str, Any]) -> List[str]:
    if change.get("directness") != "direct":
        return []
    scope = change.get("scope")
    object_ref = change.get("object_ref")
    if not object_ref:
        return []
    if scope == "cell":
        return [object_ref]
    if scope == "name":
        return [_name_node_id(object_ref)]
    if scope == "table":
        return [_table_node_id(object_ref)]
    return [object_ref]


def _ordered_graph_roots(changes: Sequence[Dict[str, Any]]) -> List[str]:
    roots: List[str] = []
    seen: Set[str] = set()
    for change in changes:
        if change.get("directness") != "direct":
            continue
        for root in _change_graph_root_refs(change):
            if root in seen:
                continue
            seen.add(root)
            roots.append(root)
    return roots


def _name_node_id(name: Any) -> str:
    return f"name:{str(name).upper()}"


def _table_node_id(name: Any) -> str:
    return f"table:{str(name).upper()}"


def _is_cell_ref(ref: str) -> bool:
    if "!" not in ref or ":" in ref or ref.startswith("["):
        return False
    try:
        from openpyxl.utils.cell import coordinate_to_tuple

        _, address = split_cell_id(ref)
        coordinate_to_tuple(address)
        return True
    except Exception:
        return False


def _is_range_ref(ref: str) -> bool:
    return "!" in ref and ":" in ref and not ref.startswith("[")


def _parse_cell_ref(ref: str) -> Optional[Tuple[str, int, int]]:
    if not _is_cell_ref(ref):
        return None
    try:
        from openpyxl.utils.cell import coordinate_to_tuple

        sheet, address = split_cell_id(ref)
        row, col = coordinate_to_tuple(address)
        return sheet, row, col
    except Exception:
        return None


def _parse_range_ref(ref: str) -> Optional[Tuple[str, int, int, int, int]]:
    if not _is_range_ref(ref):
        return None
    try:
        from openpyxl.utils.cell import range_boundaries

        sheet, cells = split_cell_id(ref)
        min_col, min_row, max_col, max_row = range_boundaries(cells.replace("$", ""))
        return sheet, min_row, max_row, min_col, max_col
    except Exception:
        return None


def _range_contains_cell(range_ref: str, cell_ref: str) -> bool:
    parsed_range = _parse_range_ref(range_ref)
    parsed_cell = _parse_cell_ref(cell_ref)
    if parsed_range is None or parsed_cell is None:
        return False
    range_sheet, min_row, max_row, min_col, max_col = parsed_range
    cell_sheet, row, col = parsed_cell
    if range_sheet != cell_sheet:
        return False
    return min_row <= row <= max_row and min_col <= col <= max_col


def _alignment_diagnostics(alignment: Dict[str, Any]) -> List[Dict[str, Any]]:
    shifted = alignment.get("summary", {}).get("shifted_semantic_matches", 0)
    if not shifted:
        return []
    return [
        {
            "severity": "info",
            "code": "STRUCTURAL_ALIGNMENT_APPLIED",
            "message": f"{shifted} cell(s) were aligned by semantic row/column identity at shifted addresses.",
        }
    ]


def _public_alignment(alignment: Dict[str, Any]) -> Dict[str, Any]:
    shifted_matches = [match for match in alignment.get("matches", []) if match.get("old_ref") != match.get("new_ref")]
    return {
        "summary": alignment.get("summary", {}),
        "shifted_matches": shifted_matches[:100],
        "shifted_matches_omitted": max(0, len(shifted_matches) - 100),
    }


def _with_workbook_ref(which: str, diagnostics: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    result = []
    for diagnostic in diagnostics:
        item = dict(diagnostic)
        item.setdefault("details", {})
        item["details"] = dict(item["details"])
        item["details"]["workbook"] = which
        result.append(item)
    return result


def _result_diagnostics(changes: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    diagnostics = []
    for change in changes:
        diagnostics.extend(change.get("warnings", []))
    return diagnostics


def _display_from_value(value: Any) -> str:
    if isinstance(value, dict) and "display_value" in value:
        return value.get("display_value", "")
    if isinstance(value, dict) and "value" in value:
        return _display_from_value(value["value"])
    return display_scalar(value)


def _signed_number(value: float, suffix: str = "") -> str:
    sign = "+" if value > 0 else ""
    if abs(value) >= 1000:
        text = f"{value:,.0f}"
    elif abs(value) >= 10:
        text = f"{value:,.1f}"
    else:
        text = f"{value:,.2f}"
    return f"{sign}{text}{suffix}"


def _signed_percent(value: float) -> str:
    sign = "+" if value > 0 else ""
    return f"{sign}{value * 100:.1f}%"


def _change_family(change: Dict[str, Any]) -> str:
    if change["kind"].startswith("formula_") or change["kind"] in {"formula_changed", "constant_to_formula", "formula_to_constant"}:
        return "formula"
    if change["kind"] == "cached_value_changed":
        return "cached-value"
    if "constant" in change["kind"] or change["kind"] in {"cell_added", "cell_deleted"}:
        return "cell"
    return change["kind"].replace("_", " ")


def _bounding_range(refs: Sequence[str]) -> str:
    from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter

    sheet_names = {split_cell_id(ref)[0] for ref in refs}
    if len(sheet_names) != 1:
        return f"{len(refs)} cells"
    sheet = next(iter(sheet_names))
    rows = []
    cols = []
    for ref in refs:
        _, address = split_cell_id(ref)
        try:
            row, col = coordinate_to_tuple(address)
            rows.append(row)
            cols.append(col)
        except Exception:
            pass
    if not rows or not cols:
        return f"{sheet}!{len(refs)} cells"
    start = f"{get_column_letter(min(cols))}{min(rows)}"
    end = f"{get_column_letter(max(cols))}{max(rows)}"
    return f"{sheet}!{start}" if start == end else f"{sheet}!{start}:{end}"


def _ref_sort_key(ref: str) -> Tuple[str, int, int, str]:
    from openpyxl.utils.cell import coordinate_to_tuple

    sheet, address = split_cell_id(ref)
    try:
        row, col = coordinate_to_tuple(address)
    except Exception:
        row, col = 0, 0
    return sheet, row, col, ref


def _best_label(labels: Sequence[Dict[str, Any]], fallback: str) -> str:
    if not labels:
        return fallback
    return max(labels, key=lambda item: item.get("confidence", 0)).get("text") or fallback


def _best_label_from_change(change: Dict[str, Any]) -> str:
    return _best_label(change.get("labels", []), change["object_ref"])


def _output_explanation(ref: str, change: Dict[str, Any], upstream_ids: Sequence[str], strength: str) -> str:
    if strength == "unexplained":
        return f"{ref} changed without a detected upstream explanation."
    if strength == "strong":
        return f"{ref} changed and has one detected upstream direct change on a dependency path; this likely explains the movement."
    if strength == "moderate":
        return f"{ref} changed and is associated with {len(upstream_ids)} upstream direct change(s)."
    return f"{ref} changed and may be related to {len(upstream_ids)} upstream direct change(s), but dependency evidence is partial."


def _config_value(config: Dict[str, Any], dotted_key: str, default: Any) -> Any:
    nested = config.get("workbook_diff", config)
    current: Any = nested
    for part in dotted_key.split("."):
        if not isinstance(current, dict) or part not in current:
            return default
        current = current[part]
    return current


def _without_hidden_sheets(snapshot: Any) -> Any:
    visible_sheets = {sheet["name"] for sheet in snapshot.sheets if sheet.get("visibility") == "visible"}
    if len(visible_sheets) == len(snapshot.sheets):
        return snapshot
    filtered_cells = {
        ref: cell
        for ref, cell in snapshot.cells.items()
        if cell.sheet_name in visible_sheets
    }
    filtered_tables = [table for table in snapshot.tables if table.get("sheet_name") in visible_sheets]
    filtered_names = []
    for name in snapshot.defined_names:
        ref = name.get("ref", "")
        if "!" not in ref or split_cell_id(ref)[0] in visible_sheets:
            filtered_names.append(name)
    diagnostics = list(snapshot.diagnostics)
    diagnostics.append(
        {
            "severity": "info",
            "code": "HIDDEN_SHEETS_IGNORED",
            "message": "Hidden and very-hidden sheets were excluded by request.",
        }
    )
    return type(snapshot)(
        schema_version=snapshot.schema_version,
        file=snapshot.file,
        sheets=[sheet for sheet in snapshot.sheets if sheet["name"] in visible_sheets],
        defined_names=filtered_names,
        tables=filtered_tables,
        external_links=snapshot.external_links,
        cells=filtered_cells,
        diagnostics=diagnostics,
    )
