from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from .formulas import VOLATILE_FUNCTIONS, formula_functions, normalize_formula, parse_formula_references
from .models import CellSnapshot, CellValue, FormulaSnapshot, WorkbookSnapshot
from .security import check_deadline, enforce_budget, inspect_workbook_package, resolve_resource_budgets
from .utils import cell_id, display_scalar, sha256_file


ASSUMPTION_SHEETS = {"ASSUMPTIONS", "INPUTS", "DRIVERS", "MODEL", "CONTROLS"}
RAW_DATA_SHEET_MARKERS = {"EXPORT", "RAW", "DATA", "STRIPE", "HUBSPOT", "CSV"}
OUTPUT_SHEETS = {"SUMMARY", "DASHBOARD", "OUTPUT", "BOARD", "KPIS", "REPORT"}
ASSUMPTION_LABEL_TERMS = {
    "growth",
    "rate",
    "churn",
    "margin",
    "tax",
    "discount",
    "cac",
    "arpu",
    "retention",
    "ltv",
    "cost",
    "price",
}
OUTPUT_LABEL_TERMS = {
    "total",
    "ltv",
    "revenue",
    "arr",
    "margin",
    "ebitda",
    "cash",
    "runway",
    "nrr",
}


def parse_workbook(
    path: Path,
    config: Optional[Dict[str, Any]] = None,
    strict: bool = False,
    options: Optional[Dict[str, Any]] = None,
    deadline: Optional[float] = None,
) -> WorkbookSnapshot:
    config = config or {}
    budgets = resolve_resource_budgets(config, options)
    package_info, diagnostics = inspect_workbook_package(path, strict=strict, budgets=budgets)
    check_deadline(deadline, f"loading {path.name}")
    formula_wb = load_workbook(path, data_only=False, keep_vba=path.suffix.lower() == ".xlsm", rich_text=False)
    value_wb = load_workbook(path, data_only=True, keep_vba=path.suffix.lower() == ".xlsm", rich_text=False)
    sheet_names = list(formula_wb.sheetnames)

    defined_names = _extract_defined_names(formula_wb)
    defined_name_map = {item["name"].upper(): item["ref"] for item in defined_names}
    tables = _extract_tables(formula_wb)
    table_map = {item["name"].upper(): item for item in tables}
    external_links = [{"part": part} for part in package_info.get("external_link_parts", [])]
    sheets = _extract_sheets(formula_wb)

    cells: Dict[str, CellSnapshot] = {}
    parsed_cell_count = 0
    formula_count = 0
    for sheet_name in sheet_names:
        check_deadline(deadline, f"parsing sheet {sheet_name}")
        formula_ws = formula_wb[sheet_name]
        value_ws = value_wb[sheet_name]
        sheet_visible = formula_ws.sheet_state == "visible"
        coordinates = set(formula_ws._cells.keys()) | set(value_ws._cells.keys())
        for row, col in sorted(coordinates):
            formula_cell = formula_ws.cell(row=row, column=col)
            value_cell = value_ws.cell(row=row, column=col)
            raw_formula_value = formula_cell.value
            cached_value = value_cell.value
            kind = _cell_kind(raw_formula_value)
            if kind == "blank" and cached_value is None and not formula_cell.comment and not formula_cell.hyperlink:
                continue
            parsed_cell_count += 1
            enforce_budget("max_parsed_cells", parsed_cell_count, budgets, f"Workbook {path.name} parsed cells")

            address = f"{get_column_letter(col)}{row}"
            ref = cell_id(sheet_name, address)
            labels = infer_labels(formula_ws, formula_cell, defined_names, tables)
            formula_snapshot = None
            value_source = cached_value if kind == "formula" else raw_formula_value
            if kind == "formula":
                formula_count += 1
                enforce_budget("max_formulas", formula_count, budgets, f"Workbook {path.name} formulas")
                formula_text = str(raw_formula_value)
                enforce_budget("max_formula_length", len(formula_text), budgets, f"Formula {ref}")
                check_deadline(deadline, f"parsing formula {ref}")
                precedents, tokens, warnings, parse_status = parse_formula_references(
                    raw_formula_value,
                    sheet_name,
                    sheet_names,
                    defined_names=defined_name_map,
                    tables=table_map,
                )
                functions = set(formula_functions(raw_formula_value))
                formula_snapshot = FormulaSnapshot(
                    raw=str(raw_formula_value),
                    normalized=normalize_formula(raw_formula_value),
                    formula_type=_formula_type(formula_cell),
                    tokens=tokens,
                    precedents=[precedent.to_dict() for precedent in precedents],
                    has_volatile_function=bool(functions & VOLATILE_FUNCTIONS),
                    has_dynamic_reference=any(function in {"INDIRECT", "OFFSET"} for function in functions),
                    has_external_reference=any(precedent.kind == "external_reference" for precedent in precedents),
                    parse_status=parse_status,
                    parse_warnings=warnings,
                )
                for warning in warnings:
                    code = "OPAQUE_DYNAMIC_REFERENCE" if "INDIRECT" in warning or "OFFSET" in warning else "FORMULA_PARSE_PARTIAL"
                    diagnostics.append(
                        {
                            "severity": "warning",
                            "code": code,
                            "message": warning,
                            "object_ref": ref,
                        }
                    )
                if cached_value is None:
                    diagnostics.append(
                        {
                            "severity": "warning",
                            "code": "FORMULA_CACHE_MISSING",
                            "message": "Formula cell has no cached value. Numeric output delta unavailable.",
                            "object_ref": ref,
                        }
                    )

            semantic_identity = infer_semantic_identity(formula_ws, formula_cell, labels, defined_names, tables)
            semantic_role = infer_semantic_role(sheet_name, kind, labels, formula_snapshot, ref, config)
            cells[ref] = CellSnapshot(
                id=ref,
                sheet_name=sheet_name,
                address=address,
                row=row,
                col=col,
                kind=kind,
                value=_cell_value(value_source, formula_cell.number_format),
                formula=formula_snapshot,
                style=_cell_style(formula_cell),
                comment=_cell_comment(formula_cell),
                hyperlink=_cell_hyperlink(formula_cell),
                labels=labels,
                semantic_id=semantic_identity.get("semantic_id"),
                semantic_identity=semantic_identity,
                semantic_role=semantic_role,
                visibility_context={
                    "sheet_visible": sheet_visible,
                    "row_hidden": bool(formula_ws.row_dimensions[row].hidden),
                    "col_hidden": bool(formula_ws.column_dimensions[get_column_letter(col)].hidden),
                },
            )

    file_info = {
        "filename": path.name,
        "sha256": sha256_file(path),
        "file_size_bytes": package_info["file_size_bytes"],
        "workbook_type": package_info["workbook_type"],
        "has_macros": package_info["has_macros"],
        "calc_mode": _calc_mode(formula_wb),
        "full_calc_on_load": _full_calc_on_load(formula_wb),
        "created_at": _iso_or_none(getattr(formula_wb.properties, "created", None)),
        "modified_at": _iso_or_none(getattr(formula_wb.properties, "modified", None)),
        "app_version": getattr(formula_wb.properties, "version", None),
        "parsed_cell_count": parsed_cell_count,
        "formula_count": formula_count,
    }
    if file_info["calc_mode"] == "manual":
        diagnostics.append(
            {
                "severity": "warning",
                "code": "MANUAL_CALCULATION_MODE",
                "message": "Workbook calculation mode is manual; cached formula values may be stale.",
            }
        )
    if file_info["full_calc_on_load"]:
        diagnostics.append(
            {
                "severity": "warning",
                "code": "FULL_CALC_ON_LOAD",
                "message": "Workbook requests full recalculation on load; cached formula values may be stale.",
            }
        )

    return WorkbookSnapshot(
        schema_version="0.1",
        file=file_info,
        sheets=sheets,
        defined_names=defined_names,
        tables=tables,
        external_links=external_links,
        cells=cells,
        diagnostics=diagnostics,
    )


def infer_labels(ws: Any, cell: Cell, defined_names: Sequence[Dict[str, Any]], tables: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    labels: List[Dict[str, Any]] = []
    ref = cell_id(ws.title, cell.coordinate)
    for name in defined_names:
        if name.get("ref") == ref:
            labels.append({"text": name["name"], "source": "defined_name", "confidence": 0.95})

    left_label = _nearest_left_label(ws, cell.row, cell.column)
    above_label = _nearest_above_label(ws, cell.row, cell.column)
    if left_label and above_label:
        labels.append(
            {
                "text": f"{above_label[0]} {left_label[0]}",
                "source": "left_neighbor+above_neighbor",
                "confidence": 0.86,
                "source_ref": f"{above_label[1]},{left_label[1]}",
            }
        )
    if left_label:
        labels.append({"text": left_label[0], "source": "left_neighbor", "confidence": 0.75, "source_ref": left_label[1]})
    if above_label:
        labels.append({"text": above_label[0], "source": "above_neighbor", "confidence": 0.65, "source_ref": above_label[1]})

    for table in tables:
        if table["sheet_name"] == ws.title and _ref_contains(table["ref"], cell.coordinate):
            labels.append({"text": table["name"], "source": "table_header", "confidence": 0.55})
            break

    labels.append({"text": ws.title, "source": "sheet_name", "confidence": 0.2})
    return _dedupe_labels(labels)


def infer_semantic_identity(
    ws: Any,
    cell: Cell,
    labels: Sequence[Dict[str, Any]],
    defined_names: Sequence[Dict[str, Any]],
    tables: Sequence[Dict[str, Any]],
) -> Dict[str, Any]:
    ref = cell_id(ws.title, cell.coordinate)
    defined_label = next((name for name in defined_names if name.get("ref") == ref), None)
    if defined_label:
        name = str(defined_label["name"])
        return {
            "semantic_id": _semantic_key(ws.title, "defined_name", [name]),
            "kind": "defined_name",
            "row_headers": [],
            "column_headers": [],
            "labels": [name],
            "source_refs": [ref],
            "confidence": 0.98,
        }

    own_label = str(cell.value).strip() if _looks_like_label(cell.value) and not str(cell.value).strip().startswith("=") else ""
    left_label = _label_by_source(labels, "left_neighbor")
    above_label = _label_by_source(labels, "above_neighbor")
    table = _table_for_cell(tables, ws.title, cell.coordinate)

    if own_label:
        axis = _label_axis(ws, cell)
        if axis == "column_header":
            return {
                "semantic_id": _semantic_key(ws.title, "column_header", [own_label]),
                "kind": "column_header",
                "row_headers": [],
                "column_headers": [own_label],
                "labels": [own_label],
                "source_refs": [ref],
                "confidence": 0.78,
            }
        return {
            "semantic_id": _semantic_key(ws.title, "row_header", [own_label]),
            "kind": "row_header",
            "row_headers": [own_label],
            "column_headers": [],
            "labels": [own_label],
            "source_refs": [ref],
            "confidence": 0.78,
        }

    if left_label and above_label:
        return {
            "semantic_id": _semantic_key(ws.title, "intersection", [left_label["text"], above_label["text"]]),
            "kind": "row_column_intersection",
            "row_headers": [left_label["text"]],
            "column_headers": [above_label["text"]],
            "labels": [left_label["text"], above_label["text"]],
            "source_refs": [left_label.get("source_ref"), above_label.get("source_ref")],
            "confidence": 0.9,
        }

    if table:
        return {
            "semantic_id": _semantic_key(ws.title, "table_cell", [table["name"], cell.coordinate]),
            "kind": "table_cell",
            "row_headers": [],
            "column_headers": [],
            "labels": [table["name"]],
            "source_refs": [table["ref"]],
            "confidence": 0.62,
        }

    if left_label:
        return {
            "semantic_id": _semantic_key(ws.title, "row_value", [left_label["text"]]),
            "kind": "row_labeled_cell",
            "row_headers": [left_label["text"]],
            "column_headers": [],
            "labels": [left_label["text"]],
            "source_refs": [left_label.get("source_ref")],
            "confidence": 0.7,
        }

    if above_label:
        return {
            "semantic_id": _semantic_key(ws.title, "column_value", [above_label["text"]]),
            "kind": "column_labeled_cell",
            "row_headers": [],
            "column_headers": [above_label["text"]],
            "labels": [above_label["text"]],
            "source_refs": [above_label.get("source_ref")],
            "confidence": 0.62,
        }

    return {
        "semantic_id": None,
        "kind": "address_only",
        "row_headers": [],
        "column_headers": [],
        "labels": [],
        "source_refs": [],
        "confidence": 0.0,
    }


def infer_semantic_role(
    sheet_name: str,
    kind: str,
    labels: Sequence[Dict[str, Any]],
    formula: Optional[FormulaSnapshot],
    ref: str,
    config: Dict[str, Any],
) -> str:
    configured_outputs = _configured_refs(config, "outputs")
    assumption_ranges = config.get("assumption_ranges", []) or config.get("workbook_diff", {}).get("assumption_ranges", [])
    raw_data_sheets = set(config.get("raw_data_sheets", []) or config.get("workbook_diff", {}).get("raw_data_sheets", []))
    if ref in configured_outputs:
        return "output"
    if _ref_in_ranges(ref, assumption_ranges):
        return "assumption"
    upper_sheet = sheet_name.upper()
    label_text = " ".join(label.get("text", "") for label in labels).lower()
    if sheet_name in raw_data_sheets or any(marker in upper_sheet for marker in RAW_DATA_SHEET_MARKERS):
        return "raw_data" if kind != "formula" else "intermediate_calculation"
    if kind == "constant":
        if upper_sheet in ASSUMPTION_SHEETS or any(term in label_text for term in ASSUMPTION_LABEL_TERMS):
            return "assumption"
        if any(label.get("source") in {"left_neighbor", "above_neighbor", "left_neighbor+above_neighbor"} for label in labels):
            return "raw_data" if "date" in label_text or "id" in label_text else "unknown"
    if kind == "formula":
        if upper_sheet in OUTPUT_SHEETS or any(term in label_text for term in OUTPUT_LABEL_TERMS):
            return "output"
        if formula and formula.has_dynamic_reference:
            return "intermediate_calculation"
        return "intermediate_calculation"
    if kind == "blank":
        return "unknown"
    return "unknown"


def _extract_sheets(wb: Any) -> List[Dict[str, Any]]:
    result = []
    for index, ws in enumerate(wb.worksheets):
        coordinates = set(ws._cells.keys())
        if coordinates:
            rows = [row for row, _ in coordinates]
            cols = [col for _, col in coordinates]
            min_row, max_row, min_col, max_col = min(rows), max(rows), min(cols), max(cols)
        else:
            min_row = max_row = min_col = max_col = 0
        result.append(
            {
                "sheet_id": str(index + 1),
                "name": ws.title,
                "index": index,
                "visibility": "very_hidden" if ws.sheet_state == "veryHidden" else ws.sheet_state,
                "dimensions": {"min_row": min_row, "max_row": max_row, "min_col": min_col, "max_col": max_col},
                "fingerprint": _sheet_fingerprint(ws),
            }
        )
    return result


def _extract_defined_names(wb: Any) -> List[Dict[str, Any]]:
    names: List[Dict[str, Any]] = []
    defined_names = getattr(wb, "defined_names", None)
    if defined_names is None:
        return names
    items: Iterable[Tuple[str, Any]]
    if hasattr(defined_names, "items"):
        items = list(defined_names.items())
    else:
        items = [(item.name, item) for item in getattr(defined_names, "definedName", [])]
    for name, defined_name in items:
        destinations = []
        try:
            destinations = list(defined_name.destinations)
        except Exception:
            destinations = []
        if destinations:
            for sheet_name, ref in destinations:
                clean_ref = ref.replace("$", "")
                names.append({"name": name, "scope": "workbook", "ref": cell_id(sheet_name, clean_ref), "raw": defined_name.attr_text})
        else:
            raw = getattr(defined_name, "attr_text", None) or str(defined_name)
            names.append({"name": name, "scope": "workbook", "ref": raw.replace("$", ""), "raw": raw})
    return names


def _extract_tables(wb: Any) -> List[Dict[str, Any]]:
    result = []
    for ws in wb.worksheets:
        table_values = ws.tables.values() if hasattr(ws.tables, "values") else []
        for table in table_values:
            result.append({"name": table.name, "sheet_name": ws.title, "ref": cell_id(ws.title, table.ref.replace("$", ""))})
    return result


def _cell_kind(raw_formula_value: Any) -> str:
    if isinstance(raw_formula_value, str) and raw_formula_value.startswith("="):
        return "formula"
    if raw_formula_value is None:
        return "blank"
    return "constant"


def _cell_value(value: Any, number_format: Optional[str]) -> CellValue:
    value_type = "blank"
    error_code = None
    if value is None:
        value_type = "blank"
    elif isinstance(value, bool):
        value_type = "boolean"
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        value_type = "number"
    elif hasattr(value, "isoformat") and not isinstance(value, str):
        value_type = "date"
    elif isinstance(value, str) and value.startswith("#"):
        value_type = "error"
        error_code = value
    else:
        value_type = "string"
    return CellValue(
        raw=value,
        typed_value=value,
        display_value=display_scalar(value),
        number_format=number_format,
        value_type=value_type,
        error_code=error_code,
    )


def _cell_style(cell: Cell) -> Dict[str, Any]:
    return {
        "style_id": getattr(cell, "style_id", None),
        "number_format": cell.number_format,
    }


def _cell_comment(cell: Cell) -> Optional[Dict[str, Any]]:
    if not cell.comment:
        return None
    return {"text": cell.comment.text, "author": cell.comment.author}


def _cell_hyperlink(cell: Cell) -> Optional[Dict[str, Any]]:
    if not cell.hyperlink:
        return None
    return {"target": cell.hyperlink.target, "location": cell.hyperlink.location, "display": cell.hyperlink.display}


def _formula_type(cell: Cell) -> str:
    data_type = getattr(cell, "data_type", None)
    if data_type == "f":
        return "normal"
    return "normal"


def _calc_mode(wb: Any) -> str:
    calc = getattr(wb, "calculation", None)
    mode = getattr(calc, "calcMode", None)
    if mode in {"auto", "manual"}:
        return mode
    return "unknown"


def _full_calc_on_load(wb: Any) -> bool:
    calc = getattr(wb, "calculation", None)
    return bool(getattr(calc, "fullCalcOnLoad", False))


def _iso_or_none(value: Any) -> Optional[str]:
    return value.isoformat() if hasattr(value, "isoformat") else None


def _nearest_left_label(ws: Any, row: int, col: int) -> Optional[Tuple[str, str]]:
    for offset in range(1, min(col, 6)):
        candidate = _existing_cell(ws, row, col - offset)
        if candidate is None:
            continue
        if _looks_like_label(candidate.value):
            return str(candidate.value).strip(), cell_id(ws.title, candidate.coordinate)
    return None


def _nearest_above_label(ws: Any, row: int, col: int) -> Optional[Tuple[str, str]]:
    for offset in range(1, min(row, 8)):
        candidate = _existing_cell(ws, row - offset, col)
        if candidate is None:
            continue
        if _looks_like_label(candidate.value):
            return str(candidate.value).strip(), cell_id(ws.title, candidate.coordinate)
    return None


def _looks_like_label(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text:
        return False
    return len(text) <= 80 and not text.startswith("=")


def _label_by_source(labels: Sequence[Dict[str, Any]], source: str) -> Optional[Dict[str, str]]:
    for label in labels:
        if label.get("source") == source and label.get("text"):
            return {"text": str(label["text"]), "source_ref": str(label.get("source_ref", ""))}
    return None


def _table_for_cell(tables: Sequence[Dict[str, Any]], sheet_name: str, address: str) -> Optional[Dict[str, Any]]:
    for table in tables:
        if table["sheet_name"] == sheet_name and _ref_contains(table["ref"], address):
            return table
    return None


def _label_axis(ws: Any, cell: Cell) -> str:
    right_values = 0
    below_values = 0
    for offset in range(1, 6):
        right_values += int(_looks_like_value_cell(_existing_cell_value(ws, cell.row, cell.column + offset)))
        below_values += int(_looks_like_value_cell(_existing_cell_value(ws, cell.row + offset, cell.column)))
    if cell.column == 1 and right_values:
        return "row_header"
    if cell.row == 1 and below_values:
        return "column_header"
    if right_values >= below_values:
        return "row_header"
    return "column_header"


def _looks_like_value_cell(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and value.strip() and not value.startswith("="):
        return False
    return True


def _existing_cell(ws: Any, row: int, col: int) -> Optional[Cell]:
    return ws._cells.get((row, col))


def _existing_cell_value(ws: Any, row: int, col: int) -> Any:
    cell = _existing_cell(ws, row, col)
    return cell.value if cell is not None else None


def _semantic_key(sheet_name: str, kind: str, parts: Sequence[Any]) -> Optional[str]:
    normalized_parts = [_normalize_semantic_text(part) for part in parts if part is not None]
    normalized_parts = [part for part in normalized_parts if part]
    if not normalized_parts:
        return None
    sheet = _normalize_semantic_text(sheet_name)
    return "::".join(["cell", sheet, kind] + normalized_parts)


def _normalize_semantic_text(value: Any) -> str:
    text = re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower())
    return text.strip("_")[:80]


def _ref_contains(range_ref: str, address: str) -> bool:
    try:
        _, ref = range_ref.rsplit("!", 1)
        from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

        row, col = coordinate_to_tuple(address)
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        return min_row <= row <= max_row and min_col <= col <= max_col
    except Exception:
        return False


def _dedupe_labels(labels: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen = set()
    result = []
    for label in sorted(labels, key=lambda item: item.get("confidence", 0), reverse=True):
        key = (label.get("text"), label.get("source"))
        if key in seen:
            continue
        seen.add(key)
        result.append(label)
    return result


def _sheet_fingerprint(ws: Any) -> str:
    parts: List[str] = [ws.title, str(ws.max_row), str(ws.max_column)]
    sampled = 0
    for (row, col), cell in sorted(ws._cells.items()):
        if sampled >= 200:
            break
        if cell.value is not None:
            value = normalize_formula(cell.value) if isinstance(cell.value, str) and cell.value.startswith("=") else str(cell.value)
            parts.append(f"{row}:{col}:{value}")
            sampled += 1
    import hashlib

    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


def _configured_refs(config: Dict[str, Any], key: str) -> set:
    nested = config.get("workbook_diff", config)
    refs = set()
    for item in nested.get(key, []) or []:
        if isinstance(item, dict) and item.get("ref"):
            refs.add(item["ref"])
        elif isinstance(item, str):
            refs.add(item)
    return refs


def _ref_in_ranges(ref: str, ranges: Sequence[str]) -> bool:
    sheet, address = ref.rsplit("!", 1)
    for range_ref in ranges:
        if "!" not in range_ref:
            continue
        range_sheet, cells = range_ref.rsplit("!", 1)
        if range_sheet not in {sheet, "*"}:
            continue
        if cells == "*":
            return True
        try:
            from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

            row, col = coordinate_to_tuple(address)
            min_col, min_row, max_col, max_row = range_boundaries(cells)
            if min_row <= row <= max_row and min_col <= col <= max_col:
                return True
        except Exception:
            continue
    return False
