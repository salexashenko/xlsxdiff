from __future__ import annotations

import json
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName

from workbook_diff.diff import diff_workbooks
from workbook_diff.reporting import write_artifacts
from workbook_diff.security import ResourceBudgetExceeded


NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_REL)


def test_simple_assumption_impact_builds_path(tmp_path: Path) -> None:
    baseline = tmp_path / "baseline.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _make_assumption_workbook(baseline, growth=0.18, revenue=1180, summary=1180)
    _make_assumption_workbook(candidate, growth=0.22, revenue=1220, summary=1220)

    result = diff_workbooks(baseline, candidate)

    direct = _find_change(result, "Assumptions!D14")
    assert direct["kind"] == "constant_changed"
    assert direct["semantic_role"] == "assumption"
    output = next(item for item in result["top_impacted_outputs"] if item["ref"] == "Summary!G31")
    assert output["explanation_strength"] == "strong"
    assert output["delta"]["absolute_delta"] == 40
    path_nodes = output["representative_paths"][0]["nodes"]
    assert path_nodes == ["Assumptions!D14", "Revenue!G22", "Summary!G31"]
    assert result["llm_summary"]["one_sentence_summary"].startswith("Summary!G31")
    assert "Assumptions!D14" in result["llm_summary"]["one_sentence_summary"]


def test_formula_logic_change_gets_token_diff(tmp_path: Path) -> None:
    baseline = tmp_path / "formula_old.xlsx"
    candidate = tmp_path / "formula_new.xlsx"
    _make_formula_change_workbook(baseline, formula="=Revenue!F22*1.18", cached=1180)
    _make_formula_change_workbook(candidate, formula="=Revenue!F22*(1+Assumptions!D14)", cached=1220)

    result = diff_workbooks(baseline, candidate)

    change = _find_change(result, "Revenue!G22")
    assert change["kind"] == "formula_reference_changed"
    assert change["evidence"][0]["type"] == "formula_diff"
    assert change["delta"]["absolute_delta"] == 40


def test_unexplained_output_change_is_not_attributed(tmp_path: Path) -> None:
    baseline = tmp_path / "unexplained_old.xlsx"
    candidate = tmp_path / "unexplained_new.xlsx"
    _make_unexplained_workbook(baseline, cached=2)
    _make_unexplained_workbook(candidate, cached=3)

    result = diff_workbooks(baseline, candidate)

    change = _find_change(result, "Summary!G31")
    assert change["kind"] == "cached_value_changed"
    assert change["directness"] == "unexplained"
    output = next(item for item in result["top_impacted_outputs"] if item["ref"] == "Summary!G31")
    assert output["explanation_strength"] == "unexplained"


def test_indirect_formula_emits_dynamic_reference_warning(tmp_path: Path) -> None:
    baseline = tmp_path / "indirect_old.xlsx"
    candidate = tmp_path / "indirect_new.xlsx"
    for path in [baseline, candidate]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws["D10"] = "D14"
        ws["B18"] = '=INDIRECT("Assumptions!" & D10)'
        wb.save(path)
        _patch_cached_values(path, {"Summary": {"B18": 10}})

    result = diff_workbooks(baseline, candidate)

    assert any(diagnostic["code"] == "OPAQUE_DYNAMIC_REFERENCE" for diagnostic in result["diagnostics"])


def test_raw_data_refresh_is_grouped(tmp_path: Path) -> None:
    baseline = tmp_path / "raw_old.xlsx"
    candidate = tmp_path / "raw_new.xlsx"
    _make_raw_data_workbook(baseline, offset=0)
    _make_raw_data_workbook(candidate, offset=100)

    result = diff_workbooks(baseline, candidate)

    assert result["summary"]["constants_changed"] == 20
    assert any(group["group_type"] == "raw_data_update" for group in result["grouped_changes"])


def test_semantic_identity_uses_row_and_column_headers(tmp_path: Path) -> None:
    baseline = tmp_path / "headers_old.xlsx"
    candidate = tmp_path / "headers_new.xlsx"
    _make_row_column_header_workbook(baseline, arr_2027=49)
    _make_row_column_header_workbook(candidate, arr_2027=55)

    result = diff_workbooks(baseline, candidate)

    change = _find_change(result, "Summary!C2")
    assert change["kind"] == "constant_changed"
    assert change["semantic_identity"]["row_headers"] == ["Anthropic ARR"]
    assert change["semantic_identity"]["column_headers"] == ["2027E"]
    assert change["semantic_id"] == "cell::summary::intersection::anthropic_arr::2027e"
    assert _best_test_label(change) == "2027E Anthropic ARR"


def test_inserted_modeling_step_aligns_shifted_cells(tmp_path: Path) -> None:
    baseline = tmp_path / "model_old.xlsx"
    candidate = tmp_path / "model_new.xlsx"
    _make_model_step_baseline(baseline)
    _make_model_step_candidate(candidate)

    result = diff_workbooks(baseline, candidate)

    assert result["structural_alignment"]["summary"]["shifted_semantic_matches"] >= 6
    assert not any(change["kind"] == "cell_deleted" for change in result["changes"])
    inserted_refs = {change["object_ref"] for change in result["changes"] if change["kind"] == "cell_added"}
    assert {"Summary!A3", "Summary!B3", "Summary!C3"}.issubset(inserted_refs)
    assert any(group["group_type"] == "modeling_step_inserted" and "GPU Expense" in group["label"] for group in result["grouped_changes"])
    assert any(group["group_type"] == "modeling_step_inserted" for group in result["llm_summary"]["top_change_groups"])

    ebitda_change = _find_change(result, "Summary!B5")
    assert ebitda_change["kind"] == "formula_reference_changed"
    assert ebitda_change["old_ref"] == "Summary!B4"
    assert ebitda_change["new_ref"] == "Summary!B5"
    assert ebitda_change["address_changed"] is True
    assert ebitda_change["delta"]["absolute_delta"] == -10
    assert not any(change.get("old_ref") == "Summary!B3" and change.get("new_ref") == "Summary!B4" for change in result["changes"])


def test_added_forecast_year_aligns_shifted_total_column(tmp_path: Path) -> None:
    baseline = tmp_path / "forecast_old.xlsx"
    candidate = tmp_path / "forecast_new.xlsx"
    _make_forecast_year_baseline(baseline)
    _make_forecast_year_candidate(candidate)

    result = diff_workbooks(baseline, candidate)

    assert any(match["old_ref"] == "Forecast!D2" and match["new_ref"] == "Forecast!E2" for match in result["structural_alignment"]["shifted_matches"])
    assert not any(change["kind"] == "cell_deleted" and change["object_ref"] == "Forecast!D2" for change in result["changes"])
    added_refs = {change["object_ref"] for change in result["changes"] if change["kind"] == "cell_added"}
    assert {"Forecast!D1", "Forecast!D2", "Forecast!D3"}.issubset(added_refs)

    total_revenue = _find_change(result, "Forecast!E2")
    assert total_revenue["kind"] == "formula_reference_changed"
    assert total_revenue["old_ref"] == "Forecast!D2"
    assert total_revenue["new_ref"] == "Forecast!E2"
    assert total_revenue["address_changed"] is True
    assert total_revenue["delta"]["absolute_delta"] == 150


def test_row_label_rename_does_not_create_false_numeric_change(tmp_path: Path) -> None:
    baseline = tmp_path / "rename_old.xlsx"
    candidate = tmp_path / "rename_new.xlsx"
    _make_row_label_rename_workbook(baseline, row_label="Revenue")
    _make_row_label_rename_workbook(candidate, row_label="Net Revenue")

    result = diff_workbooks(baseline, candidate)

    label_change = _find_change(result, "Summary!A2")
    assert label_change["kind"] == "constant_changed"
    assert label_change["old_display"] == "Revenue"
    assert label_change["new_display"] == "Net Revenue"
    assert not any(change["object_ref"] == "Summary!B2" for change in result["changes"])
    assert result["summary"]["direct_change_count"] == 1


def test_formula_to_hardcode_override_is_reported(tmp_path: Path) -> None:
    baseline = tmp_path / "override_old.xlsx"
    candidate = tmp_path / "override_new.xlsx"
    _make_hardcode_override_baseline(baseline)
    _make_hardcode_override_candidate(candidate)

    result = diff_workbooks(baseline, candidate)

    override = _find_change(result, "Summary!B4")
    assert override["kind"] == "formula_to_constant"
    assert override["semantic_role"] == "output"
    assert override["old"]["formula"]["raw"] == "=SUM(B2:B3)"
    assert override["new"]["value"]["typed_value"] == 325
    assert override["delta"]["absolute_delta"] == 25
    assert any(change["id"] == override["id"] for change in result["top_direct_changes"])


def test_hidden_sheets_can_be_excluded(tmp_path: Path) -> None:
    baseline = tmp_path / "hidden_old.xlsx"
    candidate = tmp_path / "hidden_new.xlsx"
    _make_hidden_sheet_workbook(baseline, hidden_value=10)
    _make_hidden_sheet_workbook(candidate, hidden_value=20)

    included = diff_workbooks(baseline, candidate)
    excluded = diff_workbooks(baseline, candidate, options={"include_hidden_sheets": False})

    assert _find_change(included, "HiddenCalc!B2")["kind"] == "constant_changed"
    assert excluded["changes"] == []
    assert any(diagnostic["code"] == "HIDDEN_SHEETS_IGNORED" for diagnostic in excluded["diagnostics"])


def test_style_and_comment_changes_are_opt_in(tmp_path: Path) -> None:
    baseline = tmp_path / "style_old.xlsx"
    candidate = tmp_path / "style_new.xlsx"
    _make_style_comment_workbook(baseline, styled=False)
    _make_style_comment_workbook(candidate, styled=True)

    default = diff_workbooks(baseline, candidate)
    with_options = diff_workbooks(baseline, candidate, options={"include_style_changes": True, "include_comments": True})

    assert not any(change["kind"] in {"style_changed", "comment_changed"} for change in default["changes"])
    kinds = {change["kind"] for change in with_options["changes"]}
    assert {"style_changed", "comment_changed"}.issubset(kinds)


def test_external_workbook_formula_reference_change_is_classified(tmp_path: Path) -> None:
    baseline = tmp_path / "external_old.xlsx"
    candidate = tmp_path / "external_new.xlsx"
    _make_external_reference_workbook(baseline, "='[budget_old.xlsx]Inputs'!A1", cached=10)
    _make_external_reference_workbook(candidate, "='[budget_new.xlsx]Inputs'!A1", cached=12)

    result = diff_workbooks(baseline, candidate)

    change = _find_change(result, "Summary!B2")
    assert change["kind"] == "formula_external_reference_changed"
    assert change["delta"]["absolute_delta"] == 2
    assert any(warning["type"] == "formula_diff" and warning["token_diff"][0]["type"] == "reference" for warning in change["evidence"])


def test_named_range_change_propagates_as_non_cell_root(tmp_path: Path) -> None:
    baseline = tmp_path / "name_old.xlsx"
    candidate = tmp_path / "name_new.xlsx"
    _make_named_range_driver_workbook(baseline, target_row=13, cached=110)
    _make_named_range_driver_workbook(candidate, target_row=14, cached=120)

    result = diff_workbooks(baseline, candidate)

    name_change = _find_change(result, "Growth_2027")
    assert name_change["kind"] == "defined_name_changed"
    output_change = _find_change(result, "Summary!B2")
    assert output_change["directness"] == "propagated"
    output = next(item for item in result["top_impacted_outputs"] if item["ref"] == "Summary!B2")
    assert output["explanation_strength"] == "moderate"
    assert output["upstream_change_ids"] == [name_change["id"]]
    assert any(path["nodes"][0] == "name:GROWTH_2027" for path in output["representative_paths"])
    assert output["representative_paths"][0]["display_nodes"][0]["ref"] == "Growth_2027"
    assert any(factor["code"] == "non_cell_root" for factor in output["confidence_factors"])
    assert any(group["group_type"] == "named_range_rebound" for group in result["grouped_changes"])
    assert [change["id"] for change in result["llm_summary"]["top_direct_changes"]] == [name_change["id"]]
    assert "named range Growth_2027 rebinding" in result["llm_summary"]["one_sentence_summary"]
    assert not any(change["ref"] == "Assumptions!D14" for change in result["llm_summary"]["top_direct_changes"])


def test_changed_cell_inside_large_range_reaches_output(tmp_path: Path) -> None:
    baseline = tmp_path / "large_range_old.xlsx"
    candidate = tmp_path / "large_range_new.xlsx"
    _make_large_range_workbook(baseline, raw_value=10, cached=10)
    _make_large_range_workbook(candidate, raw_value=15, cached=15)

    result = diff_workbooks(baseline, candidate, options={"max_range_expand_cells": 100})

    output_change = _find_change(result, "Summary!B2")
    assert output_change["directness"] == "propagated"
    output = next(item for item in result["top_impacted_outputs"] if item["ref"] == "Summary!B2")
    assert output["explanation_strength"] == "strong"
    assert output["representative_paths"][0]["nodes"] == ["RawData!C582", "RawData!A1:Z10000", "Summary!B2"]
    assert result["change_impact_dag"]["range_membership_summary"][0]["range_ref"] == "RawData!A1:Z10000"
    assert result["change_impact_dag"]["range_membership_summary"][0]["changed_cell_count"] == 1


def test_manual_calculation_downgrades_output_confidence(tmp_path: Path) -> None:
    baseline = tmp_path / "manual_old.xlsx"
    candidate = tmp_path / "manual_new.xlsx"
    _make_manual_calc_assumption_workbook(baseline, growth=0.18, revenue=1180, summary=1180)
    _make_manual_calc_assumption_workbook(candidate, growth=0.22, revenue=1220, summary=1220)

    result = diff_workbooks(baseline, candidate)

    output = next(item for item in result["top_impacted_outputs"] if item["ref"] == "Summary!G31")
    assert output["explanation_strength"] == "moderate"
    assert output["dependency_confidence"] == "high"
    assert output["value_delta_confidence"] == "moderate"
    assert any(factor["code"] == "manual_calculation_mode" for factor in output["confidence_factors"])


def test_formula_change_with_unchanged_cache_downgrades_value_confidence(tmp_path: Path) -> None:
    baseline = tmp_path / "formula_cache_old.xlsx"
    candidate = tmp_path / "formula_cache_new.xlsx"
    _make_formula_change_workbook(baseline, formula="=Revenue!F22*2", cached=2000)
    _make_formula_change_workbook(candidate, formula="=Revenue!F22*3", cached=2000)

    result = diff_workbooks(baseline, candidate, config={"outputs": [{"ref": "Revenue!G22", "name": "Revenue"}]})

    output = next(item for item in result["top_impacted_outputs"] if item["ref"] == "Revenue!G22")
    assert output["dependency_confidence"] == "high"
    assert output["value_delta_confidence"] == "moderate"
    assert any(factor["code"] == "formula_changed_cache_unchanged" for factor in output["confidence_factors"])
    assert "formula changed while the cached value stayed at 2000" in result["llm_summary"]["one_sentence_summary"]


def test_resource_budget_caps_parsed_cells(tmp_path: Path) -> None:
    baseline = tmp_path / "budget_cells_old.xlsx"
    candidate = tmp_path / "budget_cells_new.xlsx"
    _make_row_column_header_workbook(baseline, arr_2027=49)
    _make_row_column_header_workbook(candidate, arr_2027=55)

    with pytest.raises(ResourceBudgetExceeded, match="max_parsed_cells"):
        diff_workbooks(baseline, candidate, options={"max_parsed_cells": 1})


def test_resource_budget_caps_formula_length(tmp_path: Path) -> None:
    baseline = tmp_path / "budget_formula_old.xlsx"
    candidate = tmp_path / "budget_formula_new.xlsx"
    _make_formula_change_workbook(baseline, formula="=Revenue!F22*2", cached=2000)
    _make_formula_change_workbook(candidate, formula="=Revenue!F22*3", cached=3000)

    with pytest.raises(ResourceBudgetExceeded, match="max_formula_length"):
        diff_workbooks(baseline, candidate, options={"max_formula_length": 5})


def test_resource_budget_caps_graph_edges(tmp_path: Path) -> None:
    baseline = tmp_path / "budget_graph_old.xlsx"
    candidate = tmp_path / "budget_graph_new.xlsx"
    _make_assumption_workbook(baseline, growth=0.18, revenue=1180, summary=1180)
    _make_assumption_workbook(candidate, growth=0.22, revenue=1220, summary=1220)

    with pytest.raises(ResourceBudgetExceeded, match="max_graph_edges"):
        diff_workbooks(baseline, candidate, options={"max_graph_edges": 1})


def test_change_dag_includes_compacted_blocks(tmp_path: Path) -> None:
    baseline = tmp_path / "chain_old.xlsx"
    candidate = tmp_path / "chain_new.xlsx"
    _make_chain_workbook(baseline, driver=10, mid1=20, mid2=25, output=30)
    _make_chain_workbook(candidate, driver=15, mid1=30, mid2=35, output=40)

    result = diff_workbooks(baseline, candidate, config={"outputs": [{"ref": "Summary!B2", "name": "Final KPI"}]})

    dag = result["change_impact_dag"]
    assert dag["collapsed_node_count"] >= 1
    assert any(group["node_type"] == "collapsed_block" for group in dag["compaction_groups"])
    assert dag["compacted_edges"]


def test_cli_artifacts_are_written(tmp_path: Path) -> None:
    baseline = tmp_path / "baseline.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    out_dir = tmp_path / "out"
    _make_assumption_workbook(baseline, growth=0.18, revenue=1180, summary=1180)
    _make_assumption_workbook(candidate, growth=0.22, revenue=1220, summary=1220)

    result = diff_workbooks(baseline, candidate)
    artifacts = write_artifacts(result, out_dir, formats=["html", "json", "md"])

    assert artifacts["diff_json"].exists()
    assert artifacts["llm_summary_json"].exists()
    assert artifacts["llm_summary_md"].exists()
    assert artifacts["html"].exists()
    assert artifacts["markdown"].exists()
    assert artifacts["changed_cells_csv"].exists()
    diff_json = json.loads(artifacts["diff_json"].read_text(encoding="utf-8"))
    assert diff_json["summary"]["direct_change_count"] >= 1
    llm_json = json.loads(artifacts["llm_summary_json"].read_text(encoding="utf-8"))
    assert llm_json["one_sentence_summary"]
    html = artifacts["html"].read_text(encoding="utf-8")
    assert "<svg" in html
    assert "Focused change impact graph" in html


def _make_assumption_workbook(path: Path, growth: float, revenue: int, summary: int) -> None:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = False
    ws = wb.active
    ws.title = "Assumptions"
    ws["A14"] = "Growth Rate"
    ws["D13"] = "2026"
    ws["D14"] = growth
    ws["D14"].number_format = "0.0%"
    rev = wb.create_sheet("Revenue")
    rev["F22"] = 1000
    rev["G21"] = "2027 Revenue"
    rev["G22"] = "=Revenue!F22*(1+Assumptions!D14)"
    customers = wb.create_sheet("Customers")
    customers["G22"] = 1
    summary_ws = wb.create_sheet("Summary")
    summary_ws["A31"] = "Total LTV"
    summary_ws["G31"] = "=Revenue!G22/Customers!G22"
    wb.save(path)
    _patch_cached_values(path, {"Revenue": {"G22": revenue}, "Summary": {"G31": summary}})


def _make_formula_change_workbook(path: Path, formula: str, cached: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    ws["D14"] = 0.22
    rev = wb.create_sheet("Revenue")
    rev["F22"] = 1000
    rev["G22"] = formula
    wb.save(path)
    _patch_cached_values(path, {"Revenue": {"G22": cached}})


def _make_unexplained_workbook(path: Path, cached: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A31"] = "Total LTV"
    ws["G31"] = "=1+1"
    wb.save(path)
    _patch_cached_values(path, {"Summary": {"G31": cached}})


def _make_raw_data_workbook(path: Path, offset: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Stripe Export"
    ws.append(["id", "amount", "status"])
    for row in range(2, 12):
        ws.cell(row=row, column=1).value = f"cus_{row}"
        ws.cell(row=row, column=2).value = row + offset
        ws.cell(row=row, column=3).value = "paid" if offset else "open"
    wb.save(path)


def _make_row_column_header_workbook(path: Path, arr_2027: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "2026E"
    ws["C1"] = "2027E"
    ws["A2"] = "Anthropic ARR"
    ws["B2"] = 30
    ws["C2"] = arr_2027
    wb.save(path)


def _make_model_step_baseline(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "2026E"
    ws["C1"] = "2027E"
    ws["A2"] = "Revenue"
    ws["B2"] = 100
    ws["C2"] = 120
    ws["A3"] = "Gross Margin"
    ws["B3"] = "=B2*0.4"
    ws["C3"] = "=C2*0.4"
    ws["A4"] = "EBITDA"
    ws["B4"] = "=B3"
    ws["C4"] = "=C3"
    wb.save(path)
    _patch_cached_values(path, {"Summary": {"B3": 40, "C3": 48, "B4": 40, "C4": 48}})


def _make_model_step_candidate(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "2026E"
    ws["C1"] = "2027E"
    ws["A2"] = "Revenue"
    ws["B2"] = 100
    ws["C2"] = 120
    ws["A3"] = "GPU Expense"
    ws["B3"] = 10
    ws["C3"] = 15
    ws["A4"] = "Gross Margin"
    ws["B4"] = "=B2*0.4"
    ws["C4"] = "=C2*0.4"
    ws["A5"] = "EBITDA"
    ws["B5"] = "=B4-B3"
    ws["C5"] = "=C4-C3"
    wb.save(path)
    _patch_cached_values(path, {"Summary": {"B4": 40, "C4": 48, "B5": 30, "C5": 33}})


def _make_forecast_year_baseline(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast"
    ws["A1"] = "Metric"
    ws["B1"] = "2026E"
    ws["C1"] = "2027E"
    ws["D1"] = "Total"
    ws["A2"] = "Revenue"
    ws["B2"] = 100
    ws["C2"] = 120
    ws["D2"] = "=SUM(B2:C2)"
    ws["A3"] = "EBITDA"
    ws["B3"] = 30
    ws["C3"] = 40
    ws["D3"] = "=SUM(B3:C3)"
    wb.save(path)
    _patch_cached_values(path, {"Forecast": {"D2": 220, "D3": 70}})


def _make_forecast_year_candidate(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast"
    ws["A1"] = "Metric"
    ws["B1"] = "2026E"
    ws["C1"] = "2027E"
    ws["D1"] = "2028E"
    ws["E1"] = "Total"
    ws["A2"] = "Revenue"
    ws["B2"] = 100
    ws["C2"] = 120
    ws["D2"] = 150
    ws["E2"] = "=SUM(B2:D2)"
    ws["A3"] = "EBITDA"
    ws["B3"] = 30
    ws["C3"] = 40
    ws["D3"] = 50
    ws["E3"] = "=SUM(B3:D3)"
    wb.save(path)
    _patch_cached_values(path, {"Forecast": {"E2": 370, "E3": 120}})


def _make_row_label_rename_workbook(path: Path, row_label: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "2027E"
    ws["A2"] = row_label
    ws["B2"] = 100
    wb.save(path)


def _make_hardcode_override_baseline(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "2027E"
    ws["A2"] = "Revenue"
    ws["B2"] = 100
    ws["A3"] = "Services Revenue"
    ws["B3"] = 200
    ws["A4"] = "Total Revenue"
    ws["B4"] = "=SUM(B2:B3)"
    wb.save(path)
    _patch_cached_values(path, {"Summary": {"B4": 300}})


def _make_hardcode_override_candidate(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "2027E"
    ws["A2"] = "Revenue"
    ws["B2"] = 100
    ws["A3"] = "Services Revenue"
    ws["B3"] = 200
    ws["A4"] = "Total Revenue"
    ws["B4"] = 325
    wb.save(path)


def _make_hidden_sheet_workbook(path: Path, hidden_value: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Visible output"
    ws["B1"] = 100
    hidden = wb.create_sheet("HiddenCalc")
    hidden.sheet_state = "hidden"
    hidden["A2"] = "Internal driver"
    hidden["B2"] = hidden_value
    wb.save(path)


def _make_style_comment_workbook(path: Path, styled: bool) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "2027E"
    ws["A2"] = "Revenue"
    ws["B2"] = 100
    if styled:
        ws["B2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        ws["B2"].comment = Comment("Reviewed by FP&A", "analyst")
    wb.save(path)


def _make_external_reference_workbook(path: Path, formula: str, cached: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "2027E"
    ws["A2"] = "External input"
    ws["B2"] = formula
    wb.save(path)
    _patch_cached_values(path, {"Summary": {"B2": cached}})


def _make_named_range_driver_workbook(path: Path, target_row: int, cached: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    ws["A13"] = "Old growth"
    ws["D13"] = 0.10
    ws["A14"] = "New growth"
    ws["D14"] = 0.20
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Metric"
    summary["B1"] = "2027E"
    summary["A2"] = "Revenue"
    summary["B2"] = "=100*(1+Growth_2027)"
    wb.defined_names.add(DefinedName("Growth_2027", attr_text=f"'Assumptions'!$D${target_row}"))
    wb.save(path)
    _patch_cached_values(path, {"Summary": {"B2": cached}})


def _make_large_range_workbook(path: Path, raw_value: int, cached: int) -> None:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = False
    ws = wb.active
    ws.title = "RawData"
    ws["A1"] = "Header"
    ws["C582"] = raw_value
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Metric"
    summary["B1"] = "Value"
    summary["A2"] = "Total raw amount"
    summary["B2"] = "=SUM(RawData!A1:Z10000)"
    wb.save(path)
    _patch_cached_values(path, {"Summary": {"B2": cached}})


def _make_manual_calc_assumption_workbook(path: Path, growth: float, revenue: int, summary: int) -> None:
    wb = Workbook()
    wb.calculation.calcMode = "manual"
    ws = wb.active
    ws.title = "Assumptions"
    ws["A14"] = "Growth Rate"
    ws["D13"] = "2026"
    ws["D14"] = growth
    rev = wb.create_sheet("Revenue")
    rev["F22"] = 1000
    rev["G22"] = "=Revenue!F22*(1+Assumptions!D14)"
    customers = wb.create_sheet("Customers")
    customers["G22"] = 1
    summary_ws = wb.create_sheet("Summary")
    summary_ws["A31"] = "Total LTV"
    summary_ws["G31"] = "=Revenue!G22/Customers!G22"
    wb.save(path)
    _patch_cached_values(path, {"Revenue": {"G22": revenue}, "Summary": {"G31": summary}})


def _make_chain_workbook(path: Path, driver: int, mid1: int, mid2: int, output: int) -> None:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = False
    assumptions = wb.active
    assumptions.title = "Assumptions"
    assumptions["A1"] = "Metric"
    assumptions["B1"] = "Value"
    assumptions["A2"] = "Driver"
    assumptions["B2"] = driver
    calc = wb.create_sheet("Calc")
    calc["A1"] = "Step"
    calc["B1"] = "Value"
    calc["A2"] = "Intermediate 1"
    calc["B2"] = "=Assumptions!B2*2"
    calc["A3"] = "Intermediate 2"
    calc["B3"] = "=Calc!B2+5"
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Metric"
    summary["B1"] = "Value"
    summary["A2"] = "Final KPI"
    summary["B2"] = "=Calc!B3+5"
    wb.save(path)
    _patch_cached_values(path, {"Calc": {"B2": mid1, "B3": mid2}, "Summary": {"B2": output}})


def _find_change(result: dict, ref: str) -> dict:
    return next(change for change in result["changes"] if change["object_ref"] == ref)


def _best_test_label(change: dict) -> str:
    return max(change.get("labels", []), key=lambda item: item.get("confidence", 0)).get("text", "")


def _patch_cached_values(path: Path, values: dict) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        entries = {info.filename: archive.read(info.filename) for info in archive.infolist()}
    sheet_paths = _sheet_paths(entries)
    for sheet_name, cell_values in values.items():
        worksheet_path = sheet_paths[sheet_name]
        root = ET.fromstring(entries[worksheet_path])
        for cell_ref, value in cell_values.items():
            cell = root.find(f".//{{{NS_MAIN}}}c[@r='{cell_ref}']")
            if cell is None:
                raise AssertionError(f"Missing cell {sheet_name}!{cell_ref}")
            cell.attrib.pop("t", None)
            value_node = cell.find(f"{{{NS_MAIN}}}v")
            if value_node is None:
                value_node = ET.Element(f"{{{NS_MAIN}}}v")
                formula_node = cell.find(f"{{{NS_MAIN}}}f")
                children = list(cell)
                if formula_node is not None:
                    cell.insert(children.index(formula_node) + 1, value_node)
                else:
                    cell.append(value_node)
            value_node.text = str(value)
        entries[worksheet_path] = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, data in entries.items():
            archive.writestr(name, data)


def _sheet_paths(entries: dict) -> dict:
    workbook = ET.fromstring(entries["xl/workbook.xml"])
    rels = ET.fromstring(entries["xl/_rels/workbook.xml.rels"])
    rel_targets = {}
    for rel in rels.findall(f"{{{NS_PKG_REL}}}Relationship"):
        target = rel.attrib["Target"]
        if target.startswith("/"):
            target = target.lstrip("/")
        else:
            target = "xl/" + target.lstrip("/")
        rel_targets[rel.attrib["Id"]] = target
    result = {}
    for sheet in workbook.findall(f".//{{{NS_MAIN}}}sheet"):
        name = sheet.attrib["name"]
        rid = sheet.attrib[f"{{{NS_REL}}}id"]
        result[name] = rel_targets[rid]
    return result
